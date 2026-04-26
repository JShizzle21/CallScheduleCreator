from __future__ import annotations
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
import re

MONTHS = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
}


def looks_like_block_date(text: str) -> bool:
    if not text:
        return False
    text = str(text).strip().upper()
    return bool(re.search(r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)", text) and "-" in text)


def _parse_pgy_label(text) -> Optional[int]:
    """Return an explicit PGY number from a cell label such as 'PGY2', 'PGY-2',
    'R2', or 'Interns', or None if the text does not look like a PGY label."""
    if text is None:
        return None
    s = str(text).strip().upper()
    if re.fullmatch(r"INTERNS?", s):
        return 1
    m = re.search(r"(?:PGY|R)[- ]?(\d)", s)
    if m:
        return int(m.group(1))
    return None


def detect_skip_rows(ws) -> set[int]:
    skip_rows = set()

    for r in range(3, ws.max_row + 1):
        date_like_count = 0
        for c in range(2, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if looks_like_block_date(val):
                date_like_count += 1

        if date_like_count >= 3:
            skip_rows.add(r)

    return skip_rows


def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _infer_year(month: int, academic_year_start: int) -> int:
    return academic_year_start if month >= 7 else academic_year_start + 1


def _to_date(v, academic_year_start: int) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip().upper()

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    m = re.search(r"\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\b\W*(\d{1,2})", s)
    if m:
        mon_abbr = m.group(1)
        day = int(m.group(2))
        month = MONTHS[mon_abbr]
        year = _infer_year(month, academic_year_start)
        return date(year, month, day)

    return None


def split_rotation_parts(raw: str) -> list[str]:
    if not raw:
        return []
    parts = [part.strip() for part in re.split(r"[\\/]", raw) if part.strip()]
    return parts or [raw.strip()]


@dataclass(frozen=True)
class Block:
    col: int
    start: date
    end: date


@dataclass(frozen=True)
class RotationSegment:
    block_col: int
    block_index: int
    resident_name: str
    rotation: str
    start: date
    end: date
    part_index: int
    total_parts: int


def load_blocks(xlsx_path: str, sheet_name: str, academic_year_start: int) -> list[Block]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    starts_raw: list[tuple[int, date]] = []
    for col in range(2, ws.max_column + 1):
        d = _to_date(ws.cell(row=2, column=col).value, academic_year_start)
        if d:
            starts_raw.append((col, d))

    starts = [(col, monday_of_week(d)) for col, d in starts_raw]
    starts.sort(key=lambda x: x[1])

    final_end = date(academic_year_start + 1, 6, 30)

    blocks: list[Block] = []
    for i, (col, start) in enumerate(starts):
        if i < len(starts) - 1:
            next_start = starts[i + 1][1]
            end = next_start - timedelta(days=1)
        else:
            end = final_end
        blocks.append(Block(col=col, start=start, end=end))

    year_start = date(academic_year_start, 7, 1)
    blocks[0] = Block(col=blocks[0].col, start=year_start, end=blocks[0].end)

    return blocks


class ExcelRotationLookup:
    def __init__(self, xlsx_path: str, sheet_name: str, academic_year_start: int):
        self.xlsx_path = xlsx_path
        self.sheet_name = sheet_name
        self.academic_year_start = academic_year_start
        self.academic_start = date(academic_year_start, 7, 1)
        self.academic_end = date(academic_year_start + 1, 6, 30)

        self.wb = load_workbook(self.xlsx_path, data_only=True)
        self.ws = self.wb[self.sheet_name]

        self.skip_rows = detect_skip_rows(self.ws)
        self.separator_pgy_labels: dict[int, Optional[int]] = {
            r: _parse_pgy_label(self.ws.cell(row=r, column=1).value)
            for r in self.skip_rows
        }
        self.blocks = load_blocks(self.xlsx_path, self.sheet_name, academic_year_start=self.academic_year_start)

        self.resident_row: dict[str, int] = {}
        for r in range(3, self.ws.max_row + 1):
            if r in self.skip_rows:
                continue
            name = self.ws.cell(row=r, column=1).value
            if name:
                self.resident_row[str(name).strip()] = r

        self.date_to_block: dict[date, tuple[int, date]] = {}
        for b in self.blocks:
            cur = b.start
            while cur <= b.end:
                self.date_to_block[cur] = (b.col, b.start)
                cur += timedelta(days=1)

        self.rotation_grid: dict[str, dict[int, str]] = {}
        for name, rr in self.resident_row.items():
            per_res = {}
            for b in self.blocks:
                v = self.ws.cell(row=rr, column=b.col).value
                per_res[b.col] = "" if v is None else str(v).strip()
            self.rotation_grid[name] = per_res

        self.rotation_by_resident_date: dict[str, dict[date, str]] = {}
        self.rotation_segments_by_resident: dict[str, list[RotationSegment]] = {}

        for name in self.resident_row.keys():
            per_day: dict[date, str] = {}
            segments: list[RotationSegment] = []
            for block_index, b in enumerate(self.blocks, start=1):
                raw = self.rotation_grid.get(name, {}).get(b.col, "")
                if not raw:
                    continue

                block_segments = self._build_rotation_segments(
                    resident_name=name,
                    block=b,
                    block_index=block_index,
                    raw_rotation=raw,
                )
                segments.extend(block_segments)

                for seg in block_segments:
                    cur = seg.start
                    while cur <= seg.end:
                        per_day[cur] = seg.rotation
                        cur += timedelta(days=1)

            self.rotation_by_resident_date[name] = per_day
            self.rotation_segments_by_resident[name] = segments

    def _build_rotation_segments(
        self,
        resident_name: str,
        block: Block,
        block_index: int,
        raw_rotation: str,
    ) -> list[RotationSegment]:
        parts = split_rotation_parts(raw_rotation)
        if not parts:
            return []

        if len(parts) == 1:
            return [
                RotationSegment(
                    block_col=block.col,
                    block_index=block_index,
                    resident_name=resident_name,
                    rotation=parts[0],
                    start=block.start,
                    end=block.end,
                    part_index=1,
                    total_parts=1,
                )
            ]

        if len(parts) != 2:
            raise ValueError(
                f"Unsupported split rotation format for {resident_name}, block {block_index}: '{raw_rotation}'"
            )

        first_rotation, second_rotation = parts

        if block.start == self.academic_start and block.start.weekday() != 0:
            first_sunday = block.start + timedelta(days=(6 - block.start.weekday()))
            first_half_end = min(block.end, first_sunday + timedelta(days=14))
        else:
            first_half_end = min(block.end, block.start + timedelta(days=13))

        second_half_start = first_half_end + timedelta(days=1)

        segments = [
            RotationSegment(
                block_col=block.col,
                block_index=block_index,
                resident_name=resident_name,
                rotation=first_rotation,
                start=block.start,
                end=first_half_end,
                part_index=1,
                total_parts=2,
            )
        ]

        if second_half_start <= block.end:
            segments.append(
                RotationSegment(
                    block_col=block.col,
                    block_index=block_index,
                    resident_name=resident_name,
                    rotation=second_rotation,
                    start=second_half_start,
                    end=block.end,
                    part_index=2,
                    total_parts=2,
                )
            )

        return segments

    def rotation_on_date(self, resident_name: str, d: date) -> Optional[str]:
        return self.rotation_by_resident_date.get(resident_name, {}).get(d)

    def rotation_segments_for_resident(self, resident_name: str) -> list[RotationSegment]:
        return list(self.rotation_segments_by_resident.get(resident_name, []))
