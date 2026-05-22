"""Streamlit GUI for the Call Schedule Creator.

Phase 5a — scaffold + upload section only. Settings, Run, and Results are
stubbed. See docs/gui_plan.md for the full spec.

Runtime notes:
- Requires streamlit >= 1.38 (pinned in requirements.txt).
- Uploaded files land in a per-session temp directory created under the
  OS temp location via tempfile.mkdtemp(prefix="CallScheduler_"). On
  Windows this is typically C:\\Users\\<user>\\AppData\\Local\\Temp\\.
  Files in the session tmpdir are deleted when the user clicks
  "Reset session"; orphaned tmpdirs from prior sessions (older than
  24 hours) are cleaned on launch.
- All GUI state lives in st.session_state. Nothing persists to disk
  except via "Save as defaults" (Phase 5b), which overwrites config.yaml.
"""

from __future__ import annotations

import logging
import os
import queue
import re
import shutil
import sys
import tempfile
import threading
import time
import traceback
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Optional

# scheduler_main lives at the project root (one level up from src/).
# Streamlit puts the script's own directory (src/) on sys.path automatically,
# but not the parent — add it explicitly so `from scheduler_main import ...`
# below works regardless of how app.py is launched.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from config import load_default_config, save_config

# Imported lazily-ish (top-level is fine — these modules already import on
# every CLI run). Bringing them in at module top keeps the run-button
# handler simple. The scheduler module installs no logging handlers at
# import; we attach our own per-run handler in _run_simulation_thread.
from scheduler_main import run_simulation
from data_bundle import load_data_bundle
from loader import load_completed_calls

TMPDIR_PREFIX = "CallScheduler_"
ORPHAN_AGE_SECONDS = 24 * 3600

# Fixed allowed sets for the two lexicographic-order fields. The user may
# reorder but not add or remove items — the scheduler requires exactly these.
ALLOWED_MC_SCORE_ORDER = {
    "errors",
    "unassigned",
    "upper_weekend_diff",
    "upper_weekday_diff",
    "pgy2_total_diff",
    "pgy3_total_diff",
    "intern_weekend_diff",
    "avoid_assignments",
    "warnings",
}
ALLOWED_PICK_RANK_ORDER = {
    "hard_diff_flag",
    "soft_diff_flag",
    "weighted_score",
}

# Keys that the Settings section exposes to the user. Path keys and SHEET_NAME
# are intentionally omitted — they're managed by the Upload section.
BEHAVIOR_KEYS = [
    # Common
    "ACADEMIC_DATE_START_STRING",
    "ACADEMIC_DATE_END_STRING",
    "PGY3_CUTOFF_DATE",
    "SIMULATION_RUNS",
    "USE_COMPLETED_CALLS",
    "INTERN_BLOCK1_WEEKDAY_CALLS",
    # Advanced
    "POST_CALL_DAYS",
    "MIN_SPACING_DAYS_STRONG",
    "MIN_SPACING_DAYS_MILD",
    "MAX_CALLS_IN_WINDOW",
    "ROLLING_WINDOW_DAYS",
    "MAX_DIFF_SOFT",
    "MAX_DIFF_HARD",
    "NIGHT_FLOAT_ROTATION_NAME",
    # Expert
    "FAIRNESS_GAP_WEIGHT",
    "SPACING_WEIGHT",
    "AVOID_WEIGHT",
    "YEAR_BIAS_WEIGHT",
    "PACE_WEIGHT",
    "LOOKAHEAD_WEIGHT",
    "PICK_CANDIDATE_RANK_ORDER",
    "MONTE_CARLO_SCORE_ORDER",
]

WEIGHT_KEYS = [
    "FAIRNESS_GAP_WEIGHT",
    "SPACING_WEIGHT",
    "AVOID_WEIGHT",
    "YEAR_BIAS_WEIGHT",
    "PACE_WEIGHT",
    "LOOKAHEAD_WEIGHT",
]


# ---------------------------------------------------------------------------
# Upload slot definitions
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class UploadSlot:
    key: str
    label: str
    saved_filename: str  # What we save the upload as inside the tmpdir.
    required: bool
    validator: Callable[[Path], Optional[str]]
    help: str


def _looks_like_date(v) -> bool:
    """Mirror excel_reader._to_date's accepted forms (loosely).

    The real flow-sheet loader parses date/datetime objects and a wide
    range of string formats including "JUL 1", "7/1/2025", and ISO. The
    upload validator must be at least as permissive — otherwise legit
    flow sheets get rejected here even though the loader would parse
    them fine.
    """
    if v is None:
        return False
    if isinstance(v, (date, datetime)):
        return True
    s = str(v).strip()
    if not s:
        return False
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            datetime.strptime(s, fmt)
            return True
        except ValueError:
            pass
    # Month-abbreviation form: "JUL 1", "Jul 01", "JUL-1", etc.
    if re.search(
        r"\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\b\W*\d{1,2}",
        s.upper(),
    ):
        return True
    return False


def _validate_flow(path: Path) -> Optional[str]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return f"Could not open workbook: {exc}"
    if "master_block_calendar" not in wb.sheetnames:
        return (
            "Flow sheet error: expected a sheet named 'master_block_calendar'. "
            f"Found: {', '.join(wb.sheetnames)}"
        )
    ws = wb["master_block_calendar"]
    # Row 2, column B+ should hold block start dates. Match the loader's
    # parsing rules (excel_reader._to_date) so any value the real loader
    # accepts also passes here.
    date_cells = [ws.cell(row=2, column=c).value for c in range(2, min(ws.max_column, 30) + 1)]
    if not any(_looks_like_date(v) for v in date_cells):
        return (
            "Flow sheet error: row 2 does not contain any parseable dates in "
            "columns B onward. Expected block start dates there."
        )
    return None


def _read_headers(path: Path) -> tuple[Optional[set[str]], Optional[str]]:
    """Return (lowercase header set, error). One of the two is None."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return None, f"Could not open workbook: {exc}"
    ws = wb.active
    headers = {
        str(ws.cell(row=1, column=c).value).strip().lower()
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value is not None
    }
    return headers, None


def _validate_required_headers(
    path: Path, required: set[str], label: str
) -> Optional[str]:
    headers, err = _read_headers(path)
    if err is not None:
        return err
    missing = {h.lower() for h in required} - headers
    if missing:
        return (
            f"{label} error: header row is missing required column(s): "
            f"{', '.join(sorted(missing))}. Found: {', '.join(sorted(headers))}."
        )
    return None


def _validate_rotation_rules(path: Path) -> Optional[str]:
    # loader.load_rotation_rules reads row['rotation_name'], row['pgy'],
    # row['preference'] (lower-cased by _normalize_header).
    return _validate_required_headers(
        path, {"rotation_name", "pgy", "preference"}, "Rotation rules"
    )


def _validate_no_call(path: Path) -> Optional[str]:
    """Upload-time shape check for the no-call days workbook.

    Verifies the workbook has 13 sheets named 'Block 1'..'Block 13' (any
    case) and that each sheet either is empty or contains at least one
    header row with First Name / Last Name / Start Date / End Date.
    Per-row content validation (last names match residents, dates inside
    the academic year) happens at run time.
    """
    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(path, data_only=True)
    except Exception as e:
        return f"No-call days error: could not open workbook ({e})."

    sheetnames_lower = {s.lower(): s for s in wb.sheetnames}
    missing = [
        f"Block {i}" for i in range(1, 14)
        if f"block {i}" not in sheetnames_lower
    ]
    if missing:
        return (
            "No-call days error: workbook is missing required sheet(s): "
            f"{', '.join(missing)}. Expected sheets 'Block 1' through "
            f"'Block 13'."
        )

    from loader import _find_no_call_header_row, _NO_CALL_REQUIRED_HEADERS

    for i in range(1, 14):
        actual = sheetnames_lower[f"block {i}"]
        ws = wb[actual]
        # Empty sheets are OK — only error if the sheet has content but
        # no findable header row at all.
        if ws.max_row <= 1:
            continue
        hdr = _find_no_call_header_row(ws, actual, search_after=0)
        if hdr is None:
            return (
                f"No-call days error: sheet '{actual}' has content but no "
                f"header row with all four columns "
                f"({', '.join(_NO_CALL_REQUIRED_HEADERS)}) was found."
            )

    return None


def _validate_holidays(path: Path) -> Optional[str]:
    """Upload-time shape check for the holiday-rotation workbook.

    Verifies the workbook can be opened, finds a 'Holiday' header in
    column A within the first 15 rows, finds a date row 1 or 2 rows
    below it, and identifies at least one holiday column to the right.
    Cross-validation against the flow sheet (resident names, every
    flow resident represented) happens at run time in load_data_bundle
    because it requires reading two files together.
    """
    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(path, data_only=True)
    except Exception as e:
        return f"Holidays error: could not open workbook ({e})."

    ws = wb.active

    from loader import _find_holiday_header_row, _find_holiday_date_row
    from loader import _parse_holiday_date_with_year_inference

    header_row = _find_holiday_header_row(ws)
    if header_row is None:
        return (
            "Holidays error: no 'Holiday' header in column A within the "
            "first 15 rows. Column A row should read 'Holiday' (or "
            "'Holidays') on the row where holiday names appear in the "
            "columns to the right."
        )

    # Identify holiday columns from the header row.
    holiday_cols = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None or str(v).strip() == "":
            continue
        if "total" in str(v).strip().lower():
            continue
        holiday_cols.append(c)

    if not holiday_cols:
        return (
            f"Holidays error: 'Holiday' header at row {header_row} but no "
            f"holiday-name columns to the right. Each holiday should have "
            f"its name in row {header_row} of its own column."
        )

    # Without academic_start/end at upload time, we can only check that at
    # least one cell in the next two rows parses as a date in *some* form.
    date_row = _find_holiday_date_row(
        ws, header_row, holiday_cols, academic_start=None, academic_end=None
    )
    if date_row is None:
        return (
            f"Holidays error: 'Holiday' header at row {header_row} but no "
            f"parseable date row in rows {header_row + 1} or "
            f"{header_row + 2}. Each holiday column should have a date "
            f"value directly below its name."
        )

    return None


def _validate_clinic(path: Path) -> Optional[str]:
    """Upload-time shape check for the clinic-days workbook.

    Verifies the workbook has 13 sheets named 'Block 1'..'Block 13' and
    that each sheet has a 'Date' column (case-insensitive) within the first
    10 rows whose cell below parses as a date. Per-row content validation
    (names match residents, dates fall inside each block, etc.) is deferred
    to load_data_bundle because it requires the flow sheet.
    """
    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(path, data_only=True)
    except Exception as e:
        return f"Clinic days error: could not open workbook ({e})."

    sheetnames_lower = {s.lower(): s for s in wb.sheetnames}
    missing = [
        f"Block {i}" for i in range(1, 14)
        if f"block {i}" not in sheetnames_lower
    ]
    if missing:
        return (
            "Clinic days error: workbook is missing required sheet(s): "
            f"{', '.join(missing)}. Expected sheets 'Block 1' through 'Block 13'."
        )

    from loader import _find_clinic_date_column
    from errors import DataValidationError as _DVE

    # _find_clinic_date_column returns None for empty block sheets and only
    # raises when a 'Date' header points at non-date data — which is what we
    # want to surface at upload time.
    for i in range(1, 14):
        actual = sheetnames_lower[f"block {i}"]
        ws = wb[actual]
        try:
            _find_clinic_date_column(ws, actual)
        except _DVE as e:
            return f"Clinic days error: {e}"

    return None


def _validate_completed(path: Path) -> Optional[str]:
    # loader.load_completed_calls finds columns by substring match
    # ('date' / 'upper' / 'intern'), so accept any header containing
    # 'date' to match.
    headers, err = _read_headers(path)
    if err is not None:
        return err
    if not any("date" in h for h in headers):
        return (
            "Completed calls error: no column whose header contains 'date'. "
            f"Found: {', '.join(sorted(headers))}."
        )
    return None


UPLOAD_SLOTS: list[UploadSlot] = [
    UploadSlot(
        key="flow",
        label="Rotation flow sheet",
        saved_filename="flow.xlsx",
        required=True,
        validator=_validate_flow,
        help="Must contain a sheet named 'master_block_calendar' with block start dates in row 2.",
    ),
    UploadSlot(
        key="rotation_rules",
        label="Rotation rules",
        saved_filename="rotation_rules.xlsx",
        required=True,
        validator=_validate_rotation_rules,
        help="Columns: rotation, PGY1, PGY2, PGY3. Values per row: ELIGIBLE, AVOID, or NO_CALL.",
    ),
    UploadSlot(
        key="no_call_days",
        label="No-call days",
        saved_filename="no_call_days.xlsx",
        required=True,
        validator=_validate_no_call,
        help="Workbook with sheets 'Block 1' through 'Block 13'. Each sheet has 'Interns' and 'Uppers' sections with columns First Name, Last Name, Start Date, End Date. Last names must match the flow sheet.",
    ),
    UploadSlot(
        key="holidays",
        label="Holidays",
        saved_filename="holidays.xlsx",
        required=True,
        validator=_validate_holidays,
        help="Single-sheet workbook. Column A row labelled 'Holiday' with holiday names in columns B+, dates 1-2 rows below, then a resident-per-row grid where each cell is the resident's rotation on that holiday (ER variants count as a call assignment).",
    ),
    UploadSlot(
        key="clinic_days",
        label="Clinic days",
        saved_filename="clinic_days.xlsx",
        required=False,
        validator=_validate_clinic,
        help="Optional. Workbook with sheets 'Block 1' through 'Block 13'. Each sheet has a 'Date' column (within the first 10 rows) and resident names listed to the right of each date.",
    ),
    UploadSlot(
        key="completed_calls",
        label="Completed calls",
        saved_filename="completed_calls.xlsx",
        required=False,
        validator=_validate_completed,
        help="Optional. Columns: date, upper, intern. Required only in partial-year mode.",
    ),
]


# ---------------------------------------------------------------------------
# Session-state management
# ---------------------------------------------------------------------------


def _cleanup_orphan_tmpdirs() -> None:
    """Delete CallScheduler_* tmpdirs older than 24 hours.

    Streamlit has no reliable shutdown hook, so prior sessions can leave
    tmpdirs behind. We only touch entries older than ORPHAN_AGE_SECONDS
    so concurrent sessions can't delete each other's state.
    """
    root = Path(tempfile.gettempdir())
    now = time.time()
    for child in root.glob(f"{TMPDIR_PREFIX}*"):
        try:
            if child.is_dir() and now - child.stat().st_mtime > ORPHAN_AGE_SECONDS:
                shutil.rmtree(child, ignore_errors=True)
        except OSError:
            pass


def _init_session_state() -> None:
    if "initialized" in st.session_state:
        return
    _cleanup_orphan_tmpdirs()
    st.session_state.tmpdir = Path(tempfile.mkdtemp(prefix=TMPDIR_PREFIX))
    # uploads[slot.key] = {"saved_path": Path, "original_name": str,
    #                       "uploaded_at": float, "error": Optional[str]}
    st.session_state.uploads = {}
    # Seed config once. Phase 5b will wire widgets to these values.
    config, _paths = load_default_config()
    st.session_state.config = config
    # Run-section state. See _start_run / _drain_queue / _render_run_section.
    st.session_state.run_state = "idle"
    st.session_state.last_log_lines = []
    st.session_state.last_progress = None
    st.session_state.last_result = None
    st.session_state.last_run_kind = None
    st.session_state.run_queue = None
    st.session_state.run_thread = None
    st.session_state.run_cancel_event = None
    st.session_state.run_error = None
    st.session_state.run_trace = None
    st.session_state.initialized = True


def _reset_session() -> None:
    """Delete the tmpdir and clear session state so the page starts fresh."""
    tmpdir = st.session_state.get("tmpdir")
    if tmpdir is not None:
        shutil.rmtree(tmpdir, ignore_errors=True)
    for key in list(st.session_state.keys()):
        del st.session_state[key]


# ---------------------------------------------------------------------------
# Upload section
# ---------------------------------------------------------------------------


def _save_upload(slot: UploadSlot, uploaded_file) -> None:
    """Persist the Streamlit UploadedFile to the session tmpdir and validate."""
    dest = st.session_state.tmpdir / slot.saved_filename
    with open(dest, "wb") as fh:
        fh.write(uploaded_file.getbuffer())
    error = slot.validator(dest)
    st.session_state.uploads[slot.key] = {
        "saved_path": dest,
        "original_name": uploaded_file.name,
        "uploaded_at": time.time(),
        "error": error,
    }


def _remove_upload(slot_key: str) -> None:
    entry = st.session_state.uploads.pop(slot_key, None)
    if entry is not None:
        try:
            Path(entry["saved_path"]).unlink(missing_ok=True)
        except OSError:
            pass


def _render_upload_slot(slot: UploadSlot) -> None:
    required_tag = "" if slot.required else " _(optional)_"
    st.markdown(f"**{slot.label}**{required_tag}")
    st.caption(slot.help)

    entry = st.session_state.uploads.get(slot.key)
    if entry is not None:
        cols = st.columns([4, 1])
        if entry["error"] is None:
            cols[0].success(
                f"✓ {entry['original_name']} — saved as `{slot.saved_filename}`"
            )
        else:
            cols[0].error(f"✗ {entry['original_name']} — {entry['error']}")
        if cols[1].button("Remove", key=f"remove_{slot.key}"):
            _remove_upload(slot.key)
            st.rerun()
    else:
        st.caption("No file loaded.")

    uploaded = st.file_uploader(
        f"Upload {slot.label}",
        type=["xlsx"],
        key=f"uploader_{slot.key}",
        label_visibility="collapsed",
    )
    if uploaded is not None:
        existing = entry is not None and entry["original_name"] == uploaded.name
        if not existing:
            _save_upload(slot, uploaded)
            st.rerun()

    st.divider()


def _all_required_valid() -> bool:
    for slot in UPLOAD_SLOTS:
        if not slot.required:
            continue
        entry = st.session_state.uploads.get(slot.key)
        if entry is None or entry["error"] is not None:
            return False
    return True


# ---------------------------------------------------------------------------
# Settings — validation
# ---------------------------------------------------------------------------


def _parse_date_or_none(raw) -> Optional[date]:
    """Return a date for a string or date; None if raw is empty/unparseable."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, date):
        return raw
    try:
        return date.fromisoformat(str(raw))
    except ValueError:
        return None


def _is_blank_date(raw) -> bool:
    return raw is None or raw == ""


def _compute_hard_errors(cfg: dict) -> dict[str, list[str]]:
    """Return {key: [messages]} for any hard validation failures.

    Hard errors block the Run button (see §4.6 of docs/gui_plan.md).
    """
    errs: dict[str, list[str]] = defaultdict(list)

    start_raw = cfg.get("ACADEMIC_DATE_START_STRING", "")
    end_raw = cfg.get("ACADEMIC_DATE_END_STRING", "")
    cutoff_raw = cfg.get("PGY3_CUTOFF_DATE", "")

    start = _parse_date_or_none(start_raw)
    end = _parse_date_or_none(end_raw)
    cutoff = _parse_date_or_none(cutoff_raw)

    if start is None:
        errs["ACADEMIC_DATE_START_STRING"].append("Academic year start is not a valid date.")
    if end is None:
        errs["ACADEMIC_DATE_END_STRING"].append("Academic year end is not a valid date.")
    if start and end and end <= start:
        errs["ACADEMIC_DATE_END_STRING"].append(
            "Academic year end must be after academic year start."
        )

    if not _is_blank_date(cutoff_raw):
        if cutoff is None:
            errs["PGY3_CUTOFF_DATE"].append("PGY3 cutoff is not a valid date.")
        elif start and end and not (start <= cutoff <= end):
            errs["PGY3_CUTOFF_DATE"].append(
                "PGY3 cutoff must be between the academic start and end dates "
                "(leave blank to disable)."
            )

    if int(cfg.get("SIMULATION_RUNS", 0)) < 1:
        errs["SIMULATION_RUNS"].append("Simulation runs must be at least 1.")

    for key in (
        "POST_CALL_DAYS",
        "MIN_SPACING_DAYS_STRONG",
        "MIN_SPACING_DAYS_MILD",
        "MAX_CALLS_IN_WINDOW",
        "ROLLING_WINDOW_DAYS",
        "MAX_DIFF_SOFT",
        "MAX_DIFF_HARD",
    ):
        if int(cfg.get(key, 0)) < 0:
            errs[key].append(f"{key} cannot be negative.")

    for key in WEIGHT_KEYS:
        if float(cfg.get(key, 0)) < 0:
            errs[key].append(f"{key} cannot be negative.")

    if int(cfg.get("MAX_DIFF_HARD", 0)) < int(cfg.get("MAX_DIFF_SOFT", 0)):
        errs["MAX_DIFF_HARD"].append("Hard threshold must be ≥ soft threshold.")

    if int(cfg.get("MIN_SPACING_DAYS_MILD", 0)) < int(cfg.get("MIN_SPACING_DAYS_STRONG", 0)):
        errs["MIN_SPACING_DAYS_MILD"].append("Mild spacing must be ≥ strong spacing.")

    if not str(cfg.get("NIGHT_FLOAT_ROTATION_NAME", "")).strip():
        errs["NIGHT_FLOAT_ROTATION_NAME"].append("Night Float rotation name cannot be blank.")

    mc = cfg.get("MONTE_CARLO_SCORE_ORDER")
    if not isinstance(mc, list) or set(mc) != ALLOWED_MC_SCORE_ORDER or len(mc) != len(set(mc)):
        errs["MONTE_CARLO_SCORE_ORDER"].append(
            f"Must list exactly these items once each (any order, one per line): "
            f"{', '.join(sorted(ALLOWED_MC_SCORE_ORDER))}."
        )

    pick = cfg.get("PICK_CANDIDATE_RANK_ORDER")
    if not isinstance(pick, list) or set(pick) != ALLOWED_PICK_RANK_ORDER or len(pick) != len(set(pick)):
        errs["PICK_CANDIDATE_RANK_ORDER"].append(
            f"Must list exactly these items once each (any order, one per line): "
            f"{', '.join(sorted(ALLOWED_PICK_RANK_ORDER))}."
        )

    return dict(errs)


def _compute_soft_warnings(cfg: dict) -> list[str]:
    """Return a list of soft-warning messages (§4.6). Do not block Run."""
    warnings: list[str] = []

    start = _parse_date_or_none(cfg.get("ACADEMIC_DATE_START_STRING"))
    end = _parse_date_or_none(cfg.get("ACADEMIC_DATE_END_STRING"))
    if start and end:
        duration = (end - start).days
        if duration < 300 or duration > 400:
            warnings.append(
                f"Academic year is {duration} days. Typical value is ~365."
            )
        cutoff = _parse_date_or_none(cfg.get("PGY3_CUTOFF_DATE"))
        if cutoff is not None and (end - cutoff).days > 45:
            warnings.append(
                f"PGY3s will be excluded for {(end - cutoff).days} days. "
                "That may strain PGY2 coverage."
            )

    runs = int(cfg.get("SIMULATION_RUNS", 0))
    if runs > 5000:
        warnings.append(
            "Simulation runs > 5000 — expect several minutes. "
            "Consider the Quick preview button if iterating."
        )
    elif 0 < runs < 100:
        warnings.append(
            "Low simulation run count may produce suboptimal schedules. "
            "1000 is the tuned default."
        )

    if int(cfg.get("POST_CALL_DAYS", 0)) == 0:
        warnings.append(
            "Post-call rest days = 0 — residents could be assigned on consecutive days."
        )

    if int(cfg.get("MAX_CALLS_IN_WINDOW", 0)) == 0:
        warnings.append(
            "Rolling-window cap disabled — burst patterns (multiple calls within a short span) "
            "will not be prevented."
        )

    for key in WEIGHT_KEYS:
        val = float(cfg.get(key, 0))
        if val == 0:
            warnings.append(f"{key} is 0 — this component will be ignored during candidate ranking.")
        elif val > 10:
            warnings.append(f"{key} is {val} — far above the default. This component will dominate ranking.")

    return warnings


# ---------------------------------------------------------------------------
# Settings — widget rendering
# ---------------------------------------------------------------------------


def _load_defaults_fresh() -> dict:
    """Re-read config.yaml from disk each rerun for live ● diff comparisons."""
    try:
        cfg, _paths = load_default_config()
        return cfg
    except Exception:
        return {}


def _sync_cfg_from_widgets(cfg: dict) -> None:
    """Mirror the current widget state from session_state into cfg.

    On rerun, st.session_state[w_KEY] already holds the user's latest
    widget value before our script runs. Pulling these values into cfg
    BEFORE computing the ● diff markers and hard-error list is essential:
    without this, validation and markers lag by one rerun (they'd see
    the prior rerun's final cfg, not the widget state that just changed).
    Widgets still write back to cfg during render, but that second write
    is a no-op once sync has already happened.
    """

    def _get(wk, default=None):
        return st.session_state.get(wk, default)

    start = _get("w_ACADEMIC_DATE_START_STRING")
    if isinstance(start, date):
        cfg["ACADEMIC_DATE_START_STRING"] = start.isoformat()

    end = _get("w_ACADEMIC_DATE_END_STRING")
    if isinstance(end, date):
        cfg["ACADEMIC_DATE_END_STRING"] = end.isoformat()

    # PGY3 cutoff is gated by a checkbox. If the checkbox has ever
    # rendered we trust it; otherwise keep whatever cfg already has.
    if "cb_PGY3_CUTOFF_DATE" in st.session_state:
        if st.session_state["cb_PGY3_CUTOFF_DATE"]:
            dp = st.session_state.get("dp_PGY3_CUTOFF_DATE")
            if isinstance(dp, date):
                cfg["PGY3_CUTOFF_DATE"] = dp.isoformat()
        else:
            cfg["PGY3_CUTOFF_DATE"] = ""

    for key in (
        "SIMULATION_RUNS",
        "POST_CALL_DAYS",
        "MIN_SPACING_DAYS_STRONG",
        "MIN_SPACING_DAYS_MILD",
        "MAX_CALLS_IN_WINDOW",
        "ROLLING_WINDOW_DAYS",
        "MAX_DIFF_SOFT",
        "MAX_DIFF_HARD",
    ):
        v = _get(f"w_{key}")
        if v is not None:
            cfg[key] = int(v)

    for key in ("USE_COMPLETED_CALLS", "INTERN_BLOCK1_WEEKDAY_CALLS"):
        v = _get(f"w_{key}")
        if v is not None:
            cfg[key] = int(bool(v))

    nf = _get("w_NIGHT_FLOAT_ROTATION_NAME")
    if nf is not None:
        cfg["NIGHT_FLOAT_ROTATION_NAME"] = str(nf)

    for key in WEIGHT_KEYS:
        v = _get(f"w_{key}")
        if v is not None:
            cfg[key] = float(v)

    # PICK_CANDIDATE_RANK_ORDER and MONTE_CARLO_SCORE_ORDER are no longer
    # backed by a text widget — the up/down buttons mutate cfg[key]
    # directly, so cfg is already the source of truth here.


def _diff_marker(key: str, defaults: dict) -> str:
    return " ●" if st.session_state.config.get(key) != defaults.get(key) else ""


def _label(key: str, text: str, defaults: dict) -> str:
    return f"{text}{_diff_marker(key, defaults)}"


def _render_errors(key: str, errors_by_key: dict[str, list[str]]) -> None:
    for msg in errors_by_key.get(key, []):
        st.error(msg)


def _date_input_nullable(
    label: str,
    key: str,
    value_raw: str,
    help_text: str,
    fallback_date: date,
) -> str:
    """Nullable date input. Returns ISO string, or '' if cleared.

    The checkbox gates the picker because st.date_input can't itself be
    "blank" once edited. `fallback_date` is used only on the very first
    render when cfg is empty and no prior dp_* session_state exists —
    once the user picks a date, Streamlit's session_state persists it
    across uncheck/recheck cycles.
    """
    current = _parse_date_or_none(value_raw)
    # Remember the most recent picked value so unchecking → rechecking
    # restores it. Streamlit drops dp_* session_state when the widget
    # isn't rendered, so we need our own stash.
    remember_key = f"_last_{key}"
    if current is not None:
        st.session_state[remember_key] = current

    enabled = st.checkbox(
        f"Enable {label.lower()}",
        value=current is not None,
        key=f"cb_{key}",
        help=help_text,
    )
    if not enabled:
        return ""
    initial = current or st.session_state.get(remember_key) or fallback_date
    picked = st.date_input(
        label,
        value=initial,
        key=f"dp_{key}",
        label_visibility="collapsed",
    )
    if picked:
        st.session_state[remember_key] = picked
    return picked.isoformat() if picked else ""


def _default_anchor_year() -> int:
    """Start year of the next-upcoming academic year.

    The current AY is treated as already in progress / past for scheduling
    purposes, so we anchor on the next one. Today 2026-04-23 → 2026
    (AY 2026–2027 starts Jul 1 2026). Today 2026-08-15 → 2027.
    """
    today = date.today()
    return today.year if today.month < 7 else today.year + 1


def _academic_year_options(anchor_year: int | None = None) -> list[tuple[str, date, date]]:
    """Return (label, start_date, end_date) for 10 consecutive academic years.

    Each AY runs Jul 1 of year Y → Jun 30 of year Y+1. The list starts at the
    next-upcoming AY (or `anchor_year` if given) and lists 10 forward.
    """
    if anchor_year is None:
        anchor_year = _default_anchor_year()
    options = []
    for i in range(10):
        y = anchor_year + i
        s = date(y, 7, 1)
        e = date(y + 1, 6, 30)
        label = f"{s.month}/{s.day}/{s.year} – {e.month}/{e.day}/{e.year}"
        options.append((label, s, e))
    return options


def _is_standard_academic_year(start: date | None, end: date | None) -> bool:
    """True if (start, end) is a Jul 1 → next Jun 30 pair."""
    if start is None or end is None:
        return False
    return (
        start.month == 7
        and start.day == 1
        and end.month == 6
        and end.day == 30
        and end.year == start.year + 1
    )


def _render_common_section(defaults: dict, errors: dict[str, list[str]]) -> None:
    st.markdown("**Common** — most users only edit these.")

    cfg = st.session_state.config

    start_current = _parse_date_or_none(cfg.get("ACADEMIC_DATE_START_STRING"))
    end_current = _parse_date_or_none(cfg.get("ACADEMIC_DATE_END_STRING"))

    # UI-only flag (not persisted): when OFF, show the AY dropdown; when ON,
    # show the original pair of date pickers. Default to OFF unless cfg
    # currently holds a non-standard span (then auto-flip ON so the user
    # sees their actual dates instead of a silently-rounded dropdown pick).
    if "ui_custom_ay_dates" not in st.session_state:
        st.session_state["ui_custom_ay_dates"] = not _is_standard_academic_year(
            start_current, end_current
        )

    custom_mode = st.toggle(
        "Custom academic year dates",
        value=bool(st.session_state["ui_custom_ay_dates"]),
        key="ui_custom_ay_dates",
        help="OFF: pick a standard Jul 1 – Jun 30 academic year from the list. "
        "ON: choose any custom start and end dates.",
    )

    if not custom_mode:
        # Anchor the option list on the current cfg start year so the user's
        # saved AY appears even if it's outside the today-anchored window.
        default_anchor = _default_anchor_year()
        anchor = default_anchor
        if (
            start_current is not None
            and start_current.month == 7
            and start_current.day == 1
            and start_current.year < default_anchor
        ):
            anchor = start_current.year
        options = _academic_year_options(anchor_year=anchor)
        labels = [o[0] for o in options]

        # Pick the option whose start matches cfg; else default to index 0.
        default_idx = 0
        for i, (_, s, _e) in enumerate(options):
            if start_current == s:
                default_idx = i
                break

        picked_label = st.selectbox(
            _label("ACADEMIC_DATE_START_STRING", "Academic year", defaults),
            options=labels,
            index=default_idx,
            key="w_ACADEMIC_YEAR_CHOICE",
            help="Standard Jul 1 – Jun 30 academic year.",
        )
        picked = next(o for o in options if o[0] == picked_label)
        cfg["ACADEMIC_DATE_START_STRING"] = picked[1].isoformat()
        cfg["ACADEMIC_DATE_END_STRING"] = picked[2].isoformat()
        _render_errors("ACADEMIC_DATE_START_STRING", errors)
        _render_errors("ACADEMIC_DATE_END_STRING", errors)
    else:
        start_default = start_current or date.today()
        start_picked = st.date_input(
            _label("ACADEMIC_DATE_START_STRING", "Academic year start", defaults),
            value=start_default,
            key="w_ACADEMIC_DATE_START_STRING",
            help="First day of the academic year. All scheduling begins from this date.",
        )
        cfg["ACADEMIC_DATE_START_STRING"] = start_picked.isoformat() if start_picked else ""
        _render_errors("ACADEMIC_DATE_START_STRING", errors)

        end_default = end_current or date.today()
        end_picked = st.date_input(
            _label("ACADEMIC_DATE_END_STRING", "Academic year end", defaults),
            value=end_default,
            key="w_ACADEMIC_DATE_END_STRING",
            help="Last day of the academic year (inclusive).",
        )
        cfg["ACADEMIC_DATE_END_STRING"] = end_picked.isoformat() if end_picked else ""
        _render_errors("ACADEMIC_DATE_END_STRING", errors)

    # Fallback for PGY3 cutoff when enabling from empty: 14 days before
    # academic end. Avoids defaulting to today (which is out of range).
    end_for_fallback = _parse_date_or_none(cfg.get("ACADEMIC_DATE_END_STRING"))
    pgy3_fallback = (
        end_for_fallback - timedelta(days=14) if end_for_fallback else date.today()
    )
    cfg["PGY3_CUTOFF_DATE"] = _date_input_nullable(
        _label("PGY3_CUTOFF_DATE", "PGY3 graduation cutoff", defaults),
        "PGY3_CUTOFF_DATE",
        str(cfg.get("PGY3_CUTOFF_DATE", "")),
        "PGY3s are excluded from call on this date and after. "
        "Disable to keep PGY3s eligible all year.",
        fallback_date=pgy3_fallback,
    )
    _render_errors("PGY3_CUTOFF_DATE", errors)

    cfg["SIMULATION_RUNS"] = int(
        st.number_input(
            _label("SIMULATION_RUNS", "Number of simulation runs", defaults),
            min_value=1,
            max_value=100000,
            value=int(cfg.get("SIMULATION_RUNS", 1000)),
            step=100,
            key="w_SIMULATION_RUNS",
            help="More runs = better schedule but slower. 1000 is usually sufficient.",
        )
    )
    _render_errors("SIMULATION_RUNS", errors)

    cfg["INTERN_BLOCK1_WEEKDAY_CALLS"] = int(
        st.toggle(
            _label("INTERN_BLOCK1_WEEKDAY_CALLS", "Interns take weekday calls in Block 1", defaults),
            value=bool(int(cfg.get("INTERN_BLOCK1_WEEKDAY_CALLS", 1))),
            key="w_INTERN_BLOCK1_WEEKDAY_CALLS",
            help="ON: interns cover weekday and weekend calls in Block 1. "
            "OFF: weekday calls in Block 1 are covered by Night Float.",
        )
    )
    _render_errors("INTERN_BLOCK1_WEEKDAY_CALLS", errors)

    cfg["USE_COMPLETED_CALLS"] = int(
        st.toggle(
            _label("USE_COMPLETED_CALLS", "Partial-year mode (seed from completed calls)", defaults),
            value=bool(int(cfg.get("USE_COMPLETED_CALLS", 0))),
            key="w_USE_COMPLETED_CALLS",
            help="ON: seed from completed_calls.xlsx and generate from the next day. "
            "OFF: generate a full year from scratch.",
        )
    )
    _render_errors("USE_COMPLETED_CALLS", errors)


def _render_advanced_section(defaults: dict, errors: dict[str, list[str]]) -> None:
    cfg = st.session_state.config

    def _num(key: str, label: str, help_text: str, step: int = 1, min_value: int = 0):
        cfg[key] = int(
            st.number_input(
                _label(key, label, defaults),
                min_value=min_value,
                value=int(cfg.get(key, 0)),
                step=step,
                key=f"w_{key}",
                help=help_text,
            )
        )
        _render_errors(key, errors)

    _num("POST_CALL_DAYS", "Post-call rest days",
         "Minimum days off after a call before another can be assigned. Hard constraint.")
    _num("MIN_SPACING_DAYS_STRONG", "Strong spacing threshold (days)",
         "Calls within this many days incur a strong penalty (soft — not hard-blocked).",
         min_value=1)
    _num("MIN_SPACING_DAYS_MILD", "Mild spacing threshold (days)",
         "Calls within this many days incur a mild penalty.", min_value=1)
    _num("MAX_CALLS_IN_WINDOW", "Max calls per rolling window",
         "Rolling cap: no more than this many calls per N days. Set to 0 to disable.")
    _num("ROLLING_WINDOW_DAYS", "Rolling window (days)",
         "Size of the rolling window for the cap above.", min_value=1)


def _render_expert_section(defaults: dict, errors: dict[str, list[str]]) -> None:
    cfg = st.session_state.config

    st.warning(
        "⚠️ These values control how the scheduler ranks candidates. They have "
        "been tuned through Monte Carlo testing. Modifying them may produce "
        "worse schedules. Change only if you understand the weighted-score "
        "system — see README §Scoring."
    )

    def _num(key: str, label: str, help_text: str, min_value: int = 0):
        cfg[key] = int(
            st.number_input(
                _label(key, label, defaults),
                min_value=min_value,
                value=int(cfg.get(key, 0)),
                step=1,
                key=f"w_{key}",
                help=help_text,
            )
        )
        _render_errors(key, errors)

    def _weight(key: str, label: str, help_text: str):
        cfg[key] = float(
            st.number_input(
                _label(key, label, defaults),
                min_value=0.0,
                value=float(cfg.get(key, 0.0)),
                step=0.25,
                format="%.2f",
                key=f"w_{key}",
                help=help_text,
            )
        )
        _render_errors(key, errors)

    _weight("FAIRNESS_GAP_WEIGHT", "Fairness gap weight",
            "How strongly to prefer residents behind in call count.")
    _weight("SPACING_WEIGHT", "Spacing weight",
            "How strongly to prefer residents with longer spacing since last call.")
    _weight("AVOID_WEIGHT", "Avoid-rotation weight",
            "Penalty for assigning call while on an AVOID rotation.")
    _weight("YEAR_BIAS_WEIGHT", "Year-bias weight",
            "How strongly to front-load PGY3s and back-load PGY2s.")
    _weight("PACE_WEIGHT", "Pace weight",
            "Corrective: penalizes residents ahead of their expected call pace.")
    _weight("LOOKAHEAD_WEIGHT", "Lookahead weight",
            "Anticipatory: prefers residents with less remaining eligibility runway.")

    # Fairness thresholds + Night Float code live in Expert because they
    # affect the internal ranking logic in non-obvious ways.
    _num("MAX_DIFF_SOFT", "Soft fairness threshold",
         "Call-count gap that triggers the soft fairness flag in ranking.", min_value=1)
    _num("MAX_DIFF_HARD", "Hard fairness threshold",
         "Call-count gap that triggers the hard fairness flag.", min_value=1)

    cfg["NIGHT_FLOAT_ROTATION_NAME"] = st.text_input(
        _label("NIGHT_FLOAT_ROTATION_NAME", "Night Float rotation code", defaults),
        value=str(cfg.get("NIGHT_FLOAT_ROTATION_NAME", "NF")),
        key="w_NIGHT_FLOAT_ROTATION_NAME",
        help="The exact name used for Night Float in the flow sheet (e.g. 'NF').",
    )
    _render_errors("NIGHT_FLOAT_ROTATION_NAME", errors)

    _render_rank_order(
        key="PICK_CANDIDATE_RANK_ORDER",
        label="Candidate rank order",
        help_text=(
            "Order in which candidate ranking criteria are applied "
            "(lexicographic, one per line). All items required."
        ),
        allowed=ALLOWED_PICK_RANK_ORDER,
        defaults=defaults,
        errors=errors,
    )
    _render_rank_order(
        key="MONTE_CARLO_SCORE_ORDER",
        label="Monte Carlo score order",
        help_text=(
            "Order in which schedule-level metrics are prioritized when selecting "
            "the best Monte Carlo run (one per line). All items required."
        ),
        allowed=ALLOWED_MC_SCORE_ORDER,
        defaults=defaults,
        errors=errors,
    )


def _render_rank_order(
    key: str,
    label: str,
    help_text: str,
    allowed: set[str],
    defaults: dict,
    errors: dict[str, list[str]],
) -> None:
    """Reorderable list with ▲/▼ buttons. Set is fixed (no add/remove).

    Free-text entry was too error-prone — typos silently broke the
    scheduler. Buttons enforce the invariant: cfg[key] is always a
    permutation of `allowed`. If cfg arrives malformed (extra/missing
    items, e.g. from a stale config.yaml), we show a self-heal button
    instead of trying to render reorder rows for a broken list.
    """
    cfg = st.session_state.config
    current = list(cfg.get(key) or [])

    st.markdown(f"**{label}**{_diff_marker(key, defaults)}")
    if help_text:
        st.caption(help_text)

    # Self-heal path: list is malformed. Offer a one-click reset to the
    # default order so the user isn't stuck. We don't auto-repair to
    # avoid silently mutating Saved values.
    if set(current) != allowed or len(current) != len(set(current)):
        st.error(
            f"List is invalid (must contain each of these once): "
            f"{', '.join(sorted(allowed))}. Current: {current}"
        )
        if st.button(f"Reset {label} to default order", key=f"reset_{key}"):
            default_list = defaults.get(key)
            if isinstance(default_list, list) and set(default_list) == allowed:
                cfg[key] = list(default_list)
            else:
                cfg[key] = sorted(allowed)
            st.rerun()
        _render_errors(key, errors)
        return

    n = len(current)
    for i, item in enumerate(current):
        cols = st.columns([6, 1, 1])
        cols[0].markdown(f"`{i + 1}.` {item}")
        if cols[1].button("▲", key=f"up_{key}_{i}", disabled=(i == 0),
                          help="Move up"):
            current[i - 1], current[i] = current[i], current[i - 1]
            cfg[key] = current
            st.rerun()
        if cols[2].button("▼", key=f"down_{key}_{i}", disabled=(i == n - 1),
                          help="Move down"):
            current[i + 1], current[i] = current[i], current[i + 1]
            cfg[key] = current
            st.rerun()
    _render_errors(key, errors)


# ---------------------------------------------------------------------------
# Settings — header (diff count, reset, save)
# ---------------------------------------------------------------------------


def _count_diffs(defaults: dict) -> int:
    return sum(
        1
        for key in BEHAVIOR_KEYS
        if st.session_state.config.get(key) != defaults.get(key)
    )


_INT_WIDGET_KEYS = (
    "SIMULATION_RUNS",
    "POST_CALL_DAYS",
    "MIN_SPACING_DAYS_STRONG",
    "MIN_SPACING_DAYS_MILD",
    "MAX_CALLS_IN_WINDOW",
    "ROLLING_WINDOW_DAYS",
    "MAX_DIFF_SOFT",
    "MAX_DIFF_HARD",
)
_BOOL_WIDGET_KEYS = ("USE_COMPLETED_CALLS", "INTERN_BLOCK1_WEEKDAY_CALLS")


def _seed_widget_state_from_config(cfg: dict) -> None:
    """Write each widget's session_state key from cfg.

    Required because deleting `w_*` keys to "reset" widgets is unreliable
    in Streamlit — widgets like date_input and number_input often keep
    their browser-side value through the rerun, leaving the UI stuck on
    the user's edits even though cfg is back to defaults. Explicitly
    setting `st.session_state[wkey] = <typed default>` BEFORE the next
    rerun is the canonical fix and works for every widget type.
    """
    # Dates: ACADEMIC_DATE_START / END appear in either the dropdown
    # (standard AY) or as date pickers (custom mode). Reset to dropdown
    # mode whenever the saved span is a standard Jul 1 → Jun 30 pair.
    start_d = _parse_date_or_none(cfg.get("ACADEMIC_DATE_START_STRING"))
    end_d = _parse_date_or_none(cfg.get("ACADEMIC_DATE_END_STRING"))
    standard = _is_standard_academic_year(start_d, end_d)
    st.session_state["ui_custom_ay_dates"] = not standard
    if standard:
        # Seed the dropdown selection to the matching label.
        anchor = min(start_d.year, _default_anchor_year())
        for label, s, _e in _academic_year_options(anchor_year=anchor):
            if s == start_d:
                st.session_state["w_ACADEMIC_YEAR_CHOICE"] = label
                break
    else:
        if start_d is not None:
            st.session_state["w_ACADEMIC_DATE_START_STRING"] = start_d
        if end_d is not None:
            st.session_state["w_ACADEMIC_DATE_END_STRING"] = end_d

    # PGY3 cutoff: checkbox + (optional) date picker.
    cutoff = _parse_date_or_none(cfg.get("PGY3_CUTOFF_DATE"))
    st.session_state["cb_PGY3_CUTOFF_DATE"] = cutoff is not None
    if cutoff is not None:
        st.session_state["dp_PGY3_CUTOFF_DATE"] = cutoff

    for key in _INT_WIDGET_KEYS:
        st.session_state[f"w_{key}"] = int(cfg.get(key, 0))

    for key in _BOOL_WIDGET_KEYS:
        st.session_state[f"w_{key}"] = bool(int(cfg.get(key, 0)))

    st.session_state["w_NIGHT_FLOAT_ROTATION_NAME"] = str(
        cfg.get("NIGHT_FLOAT_ROTATION_NAME", "NF")
    )

    for key in WEIGHT_KEYS:
        st.session_state[f"w_{key}"] = float(cfg.get(key, 0.0))

    # Rank-order lists are not widget-backed (the ▲/▼ buttons mutate
    # cfg[key] directly), so nothing to seed for them.


@st.dialog("Reset to defaults?")
def _confirm_reset() -> None:
    st.write("Discard all your changes and return to the saved defaults in config.yaml?")
    cols = st.columns(2)
    if cols[0].button("Yes, reset", type="primary", key="dlg_reset_yes"):
        fresh, _ = load_default_config()
        st.session_state.config = fresh
        _seed_widget_state_from_config(fresh)
        st.rerun()
    if cols[1].button("Cancel", key="dlg_reset_no"):
        st.rerun()


@st.dialog("Save as defaults?")
def _confirm_save(values: dict) -> None:
    st.write(
        "Overwrite config.yaml so your current values become the new defaults? "
        "The baseline will be permanently replaced. Comments in config.yaml are preserved."
    )
    cols = st.columns(2)
    if cols[0].button("Yes, save", type="primary", key="dlg_save_yes"):
        try:
            save_config(values)
            st.success("Defaults saved to config.yaml.")
        except Exception as exc:
            st.error(f"Failed to save: {exc}")
        st.rerun()
    if cols[1].button("Cancel", key="dlg_save_no"):
        st.rerun()


def _render_settings_header(defaults: dict, hard_errors: dict[str, list[str]]) -> None:
    cols = st.columns([4, 1, 1])
    diffs = _count_diffs(defaults)
    n_errors = sum(len(v) for v in hard_errors.values())
    msg_bits = []
    if diffs:
        msg_bits.append(f"{diffs} value(s) differ from defaults")
    if n_errors:
        msg_bits.append(f"**{n_errors} validation error(s) must be fixed**")
    cols[0].markdown(" · ".join(msg_bits) if msg_bits else "_All values match config.yaml defaults._")
    if cols[1].button("Reset to defaults", key="btn_reset"):
        _confirm_reset()
    # Only allow Save when there are no hard errors — saving an invalid
    # config would break the CLI too.
    save_disabled = n_errors > 0
    if cols[2].button("Save as defaults", key="btn_save", disabled=save_disabled):
        _confirm_save({k: st.session_state.config[k] for k in BEHAVIOR_KEYS})


def _render_settings_section() -> tuple[bool, list[str]]:
    """Render the whole Settings expander. Returns (hard_errors_exist, soft_warnings)."""
    defaults = _load_defaults_fresh()
    # Sync widget state → cfg BEFORE computing markers and errors so both
    # reflect the user's most recent interaction, not the prior rerun's
    # final cfg. Without this, markers and inline errors lag by one rerun.
    _sync_cfg_from_widgets(st.session_state.config)
    hard_errors = _compute_hard_errors(st.session_state.config)

    _render_settings_header(defaults, hard_errors)

    st.divider()
    with st.container():
        _render_common_section(defaults, hard_errors)
    with st.expander("Advanced", expanded=False):
        _render_advanced_section(defaults, hard_errors)
    with st.expander("Expert — tuned values, modify with caution", expanded=False):
        _render_expert_section(defaults, hard_errors)

    soft = _compute_soft_warnings(st.session_state.config)
    return bool(hard_errors), soft


# ---------------------------------------------------------------------------
# Run section — background thread, progress queue, log handler
# ---------------------------------------------------------------------------

# Quick preview seed count. Hardcoded for now; empirical tuning is
# tracked in docs/gui_plan.md §10.1.
QUICK_PREVIEW_RUNS = 100

# Output directory for the three xlsx/txt artifacts. Spec §2 says these
# overwrite each run, matching CLI behavior.
OUTPUT_SUBDIR = "output"
DATA_DIR_FOR_OUTPUT = "data"

# Polling cadence while a run is in flight. Drives the time.sleep() →
# st.rerun() loop at the bottom of main(). Short enough that the log
# feels live, long enough that we don't burn CPU re-rendering. See
# decision in handoff: option (a) — sleep + rerun, no extra deps.
RUN_POLL_INTERVAL_SEC = 0.4


class _QueueLogHandler(logging.Handler):
    """Logging handler that pushes formatted records onto a queue.

    Attached to the scheduler_main logger for the duration of a run so
    every logger.info() line (per-seed scores, holiday warnings, the
    'Simulation completed' line) flows into the GUI log area. Detached
    in the thread's finally block.
    """

    def __init__(self, q: "queue.Queue[dict]") -> None:
        super().__init__(level=logging.INFO)
        self._q = q
        self.setFormatter(logging.Formatter("%(message)s"))

    def emit(self, record: logging.LogRecord) -> None:
        try:
            self._q.put_nowait({"type": "log", "line": self.format(record)})
        except Exception:
            # Never let logging failures kill the run.
            pass


def _build_paths_from_uploads() -> dict:
    """Build the paths dict run_simulation expects from the session tmpdir.

    Required-slot files are guaranteed present (the Run button is gated
    on _all_required_valid). Optional slots fall back to "" — the loader
    treats empty strings as "no file" for clinic_days and
    completed_calls.
    """
    tmpdir: Path = st.session_state.tmpdir
    paths = {
        "flow_xlsx": str(tmpdir / "flow.xlsx"),
        "sheet_name": "master_block_calendar",
        "rotation_rules_xlsx": str(tmpdir / "rotation_rules.xlsx"),
        "no_call_days_xlsx": str(tmpdir / "no_call_days.xlsx"),
        "holidays_xlsx": str(tmpdir / "holidays.xlsx"),
        "clinic_days_xlsx": "",
        "completed_calls_xlsx": "",
        "data_dir": DATA_DIR_FOR_OUTPUT,
        "output_dir": OUTPUT_SUBDIR,
    }
    clinic_entry = st.session_state.uploads.get("clinic_days")
    if clinic_entry and clinic_entry["error"] is None:
        paths["clinic_days_xlsx"] = str(tmpdir / "clinic_days.xlsx")
    completed_entry = st.session_state.uploads.get("completed_calls")
    if completed_entry and completed_entry["error"] is None:
        paths["completed_calls_xlsx"] = str(tmpdir / "completed_calls.xlsx")
    return paths


def _preflight_completed_assignments(config: dict, paths: dict) -> list:
    """Mirror scheduler_main._main()'s completed-calls preflight.

    When USE_COMPLETED_CALLS is on, we need to load completed assignments
    once before kicking off the Monte Carlo loop. block1_end is required
    to classify weekday intern entries; we get it by loading the bundle
    once with use_completed_calls=False (cheap).
    """
    if not bool(int(config.get("USE_COMPLETED_CALLS", 0))):
        return []
    if not paths.get("completed_calls_xlsx"):
        raise FileNotFoundError(
            "Partial-year mode is on but no completed_calls.xlsx was uploaded."
        )

    intern_block1 = bool(int(config.get("INTERN_BLOCK1_WEEKDAY_CALLS", 0)))
    block1_end = None
    if intern_block1:
        academic_start = date.fromisoformat(
            str(config["ACADEMIC_DATE_START_STRING"])
        )
        academic_end = date.fromisoformat(
            str(config["ACADEMIC_DATE_END_STRING"])
        )
        bundle = load_data_bundle(
            paths,
            academic_year_start=academic_start.year,
            intern_block1_weekday_calls=True,
            use_completed_calls=False,
            academic_start_date=academic_start,
            academic_end_date=academic_end,
        )
        block1_end = bundle.block1_end
    return load_completed_calls(paths["completed_calls_xlsx"], block1_end=block1_end)


def _run_simulation_thread(
    num_runs: int,
    config: dict,
    paths: dict,
    completed_assignments: list,
    q: "queue.Queue[dict]",
    cancel_event: threading.Event,
) -> None:
    """Background-thread entry point. Streams progress + logs into the queue.

    Attaches a QueueLogHandler to the scheduler_main logger so every
    logger.info() line flows into the GUI log area; detaches it in the
    finally block. Final outcome is one of: done / cancelled / error.
    """
    sched_logger = logging.getLogger("scheduler_main")
    handler = _QueueLogHandler(q)
    prev_level = sched_logger.level
    sched_logger.addHandler(handler)
    sched_logger.setLevel(logging.INFO)
    try:
        def _progress(completed: int, total: int, info: dict) -> None:
            q.put_nowait({
                "type": "progress",
                "completed": completed,
                "total": total,
                "best_seed": info.get("best_seed"),
                "best_score": info.get("best_score"),
            })

        result = run_simulation(
            num_runs=num_runs,
            config=config,
            paths=paths,
            completed_assignments=completed_assignments,
            progress_callback=_progress,
            cancel_event=cancel_event,
        )
        if result is None:
            q.put_nowait({"type": "cancelled"})
        else:
            # Write artifacts on the worker thread so download buttons can
            # stream them straight from disk in the main thread (matches
            # the bytes the CLI produces — see handoff #3).
            from scheduler_main import export_result
            try:
                export_result(result, paths=paths)
            except PermissionError as exc:
                # Windows file lock: user has call_schedule.xlsx (or one
                # of the other outputs) open in Excel, which blocks the
                # rewrite. Surface a humane fix rather than a stack trace.
                q.put_nowait({
                    "type": "error",
                    "message": (
                        "Could not write output files because one of them "
                        "is open in Excel. Close call_schedule.xlsx, "
                        "call_totals.xlsx, and audit_report.txt, then run "
                        f"again. (Underlying error: {exc})"
                    ),
                    "trace": traceback.format_exc(),
                })
                return
            except Exception as exc:
                q.put_nowait({
                    "type": "error",
                    "message": f"Schedule generated but writing output files failed: {exc}",
                    "trace": traceback.format_exc(),
                })
                return
            q.put_nowait({"type": "done", "result": result})
    except Exception as exc:
        q.put_nowait({
            "type": "error",
            "message": str(exc),
            "trace": traceback.format_exc(),
        })
    finally:
        sched_logger.removeHandler(handler)
        sched_logger.setLevel(prev_level)


def _drain_queue() -> None:
    """Pull all pending events from the run queue into session_state.

    Called at the top of every rerun while a run is in flight. Mutates
    last_log_lines, last_progress, run_state, last_result, run_error.
    """
    q: Optional["queue.Queue[dict]"] = st.session_state.get("run_queue")
    if q is None:
        return
    while True:
        try:
            evt = q.get_nowait()
        except queue.Empty:
            break
        kind = evt.get("type")
        if kind == "log":
            st.session_state.last_log_lines.append(evt["line"])
            # Cap to last 500 lines so memory + render stay bounded.
            if len(st.session_state.last_log_lines) > 500:
                del st.session_state.last_log_lines[:-500]
        elif kind == "progress":
            st.session_state.last_progress = {
                "completed": evt["completed"],
                "total": evt["total"],
                "best_seed": evt["best_seed"],
                "best_score": evt["best_score"],
            }
        elif kind == "done":
            st.session_state.last_result = evt["result"]
            st.session_state.run_state = "done"
        elif kind == "cancelled":
            st.session_state.run_state = "cancelled"
        elif kind == "error":
            st.session_state.run_error = evt["message"]
            st.session_state.run_trace = evt.get("trace", "")
            st.session_state.run_state = "error"


def _start_run(num_runs: int, kind: str) -> None:
    """Kick off a background simulation thread.

    Clears prior results/log/progress, builds paths + preflight, then
    launches the thread. State transitions to 'running'; the next rerun
    picks up progress via _drain_queue.
    """
    config = dict(st.session_state.config)
    paths = _build_paths_from_uploads()
    Path(f"{paths['data_dir']}/{paths['output_dir']}").mkdir(parents=True, exist_ok=True)

    try:
        completed = _preflight_completed_assignments(config, paths)
    except Exception as exc:
        st.session_state.run_state = "error"
        st.session_state.run_error = f"Preflight failed: {exc}"
        st.session_state.run_trace = traceback.format_exc()
        return

    st.session_state.last_result = None
    st.session_state.last_log_lines = []
    st.session_state.last_progress = {
        "completed": 0,
        "total": num_runs,
        "best_seed": None,
        "best_score": None,
    }
    st.session_state.last_run_kind = kind
    st.session_state.run_error = None
    st.session_state.run_trace = None
    st.session_state.run_queue = queue.Queue()
    st.session_state.run_cancel_event = threading.Event()
    st.session_state.run_state = "running"

    thread = threading.Thread(
        target=_run_simulation_thread,
        args=(
            num_runs,
            config,
            paths,
            completed,
            st.session_state.run_queue,
            st.session_state.run_cancel_event,
        ),
        daemon=True,
    )
    st.session_state.run_thread = thread
    thread.start()


def _request_cancel() -> None:
    ev: Optional[threading.Event] = st.session_state.get("run_cancel_event")
    if ev is not None:
        ev.set()


def _render_run_section(uploads_ready: bool, settings_has_errors: bool, soft_warnings: list[str]) -> None:
    state = st.session_state.get("run_state", "idle")

    if not uploads_ready:
        st.warning("Upload all required files above before running.")
    if settings_has_errors:
        st.error("Fix the validation errors in Settings before running.")

    # Partial-year mode requires the completed_calls upload to actually
    # be present + valid. Caught here (not in Settings) because it's a
    # cross-cut between Settings and Uploads.
    partial_mode_missing_file = False
    if bool(int(st.session_state.config.get("USE_COMPLETED_CALLS", 0))):
        completed_entry = st.session_state.uploads.get("completed_calls")
        if completed_entry is None or completed_entry.get("error") is not None:
            partial_mode_missing_file = True
            st.error(
                "Partial-year mode is ON but no valid completed_calls.xlsx is "
                "uploaded. Upload it in section 1, or turn off "
                "\"Partial-year mode\" in Settings."
            )
    if soft_warnings:
        with st.expander(f"Configuration warnings ({len(soft_warnings)})", expanded=False):
            for w in soft_warnings:
                st.warning(w)
    if state == "running":
        st.info(
            "A simulation is in progress. Avoid refreshing the browser — a "
            "hard refresh will discard the live progress (the background "
            "process keeps running but its result is lost)."
        )

    full_runs = int(st.session_state.config.get("SIMULATION_RUNS", 1000))

    cols = st.columns([1, 1, 4])
    run_disabled = (
        not uploads_ready
        or settings_has_errors
        or partial_mode_missing_file
        or state == "running"
    )
    if cols[0].button(
        f"Quick preview ({QUICK_PREVIEW_RUNS} runs)",
        disabled=run_disabled,
        key="btn_quick",
        help="Fast iteration — runs a small number of seeds. Uses the same config as Full run.",
    ):
        _start_run(QUICK_PREVIEW_RUNS, kind="quick")
        st.rerun()
    if cols[1].button(
        f"Full run ({full_runs} runs)",
        type="primary",
        disabled=run_disabled,
        key="btn_full",
    ):
        _start_run(full_runs, kind="full")
        st.rerun()
    if state == "running":
        if cols[2].button("Cancel", key="btn_cancel"):
            _request_cancel()

    if state in ("running", "done", "cancelled", "error"):
        prog = st.session_state.get("last_progress") or {}
        completed = prog.get("completed", 0)
        total = prog.get("total", 1) or 1
        fraction = min(1.0, completed / total) if total else 0.0
        best_score = prog.get("best_score")
        best_seed = prog.get("best_seed")
        if state == "running":
            label = f"Run {completed} / {total}"
        elif state == "done":
            label = f"Completed {completed} / {total}"
        elif state == "cancelled":
            label = f"Cancelled at {completed} / {total}"
        else:
            label = f"Error at {completed} / {total}"
        if best_score is not None:
            label += f" — best so far: seed={best_seed}, score={best_score}"
        st.progress(fraction, text=label)

        log_lines = st.session_state.get("last_log_lines") or []
        if log_lines:
            st.text_area(
                "Live log",
                value="\n".join(log_lines[-200:]),
                height=200,
                key="run_log_view",
                disabled=True,
            )

    if state == "error":
        st.error(st.session_state.get("run_error") or "Simulation failed.")
        with st.expander("Error details (traceback)"):
            st.code(st.session_state.get("run_trace") or "", language="text")
    elif state == "cancelled" and st.session_state.get("last_result") is None:
        st.info("Cancelled — no schedule generated yet.")


# ---------------------------------------------------------------------------
# Results section
# ---------------------------------------------------------------------------


# Map the GUI table's PGY tint colors. Mirrors the xlsx fills used by
# write_call_totals_xlsx (PGY1 green, PGY2 blue, PGY3 red) so the table
# reads the same as the spreadsheet.
_PGY_ROW_COLORS = {
    1: "#D9EAD3",
    2: "#CFE2F3",
    3: "#F4CCCC",
}


def _format_score_tuple(score, order: list[str]) -> str:
    if score is None:
        return "—"
    return ", ".join(f"{k}={v}" for k, v in zip(order, score))


def _output_paths() -> dict:
    base = Path(DATA_DIR_FOR_OUTPUT) / OUTPUT_SUBDIR
    return {
        "schedule": base / "call_schedule.xlsx",
        "totals": base / "call_totals.xlsx",
        "audit": base / "audit_report.txt",
    }


def _read_bytes(p: Path) -> bytes:
    try:
        return p.read_bytes()
    except OSError:
        return b""


def _render_status_banner(result: dict) -> None:
    audit = result["audit_data"]
    n_errors = len(audit.get("errors", []))
    n_warnings = len(audit.get("warnings", []))
    n_unassigned = len(result.get("unassigned_rows", []))
    score = audit.get("monte_carlo_score_order", [])
    # Recompute MC score from the result so the banner matches what the
    # selector picked (this run's actual tuple, not the last-seen partial).
    from scheduler_main import monte_carlo_score
    mc = monte_carlo_score(result)
    score_str = _format_score_tuple(mc, score)
    summary = (
        f"{n_warnings} warning(s), {n_errors} error(s), "
        f"{n_unassigned} unassigned slot(s). MC score: {score_str}"
    )
    if n_errors > 0:
        st.error(f"Completed with errors. {summary}")
    elif n_warnings > 0 or n_unassigned > 0:
        st.warning(f"Completed with warnings. {summary}")
    else:
        st.success(f"Schedule generated successfully. {summary}")


def _render_downloads() -> None:
    paths = _output_paths()
    cols = st.columns(3)
    cols[0].download_button(
        "⬇ call_schedule.xlsx",
        data=_read_bytes(paths["schedule"]),
        file_name="call_schedule.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=not paths["schedule"].exists(),
        key="dl_schedule",
    )
    cols[1].download_button(
        "⬇ call_totals.xlsx",
        data=_read_bytes(paths["totals"]),
        file_name="call_totals.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=not paths["totals"].exists(),
        key="dl_totals",
    )
    cols[2].download_button(
        "⬇ audit_report.txt",
        data=_read_bytes(paths["audit"]),
        file_name="audit_report.txt",
        mime="text/plain",
        disabled=not paths["audit"].exists(),
        key="dl_audit",
    )


def _render_audit_view(result: dict) -> None:
    audit = result["audit_data"]
    errors = audit.get("errors", [])
    warnings = audit.get("warnings", [])
    fairness = audit.get("fairness_summary", {})
    unassigned = result.get("unassigned_rows", [])
    avoid_assigns = audit.get("avoid_assignments", [])

    with st.expander(
        f"Errors ({len(errors)})" if errors else "Errors",
        expanded=False,
    ):
        if errors:
            for msg in errors:
                st.error(msg)
        else:
            st.success("No errors ✓")

    with st.expander(f"Warnings ({len(warnings)})", expanded=False):
        if warnings:
            for msg in warnings:
                st.warning(msg)
        else:
            st.caption("No warnings.")

    with st.expander("Fairness summary", expanded=False):
        # Group the 7 keys * 3 stats from fairness_summary into a tidy
        # min/max/diff table by metric prefix.
        rows = []
        seen = set()
        for key in fairness:
            if key.endswith("_diff") or key.endswith("_min") or key.endswith("_max"):
                base = key.rsplit("_", 1)[0]
                seen.add(base)
        for base in sorted(seen):
            rows.append({
                "metric": base,
                "min": fairness.get(f"{base}_min", ""),
                "max": fairness.get(f"{base}_max", ""),
                "diff": fairness.get(f"{base}_diff", ""),
            })
        if rows:
            st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)
        else:
            st.caption("No fairness data.")

    with st.expander(
        f"Unassigned slots ({len(unassigned)})" if unassigned else "Unassigned slots",
        expanded=False,
    ):
        if unassigned:
            df = pd.DataFrame(unassigned)
            st.dataframe(df, hide_index=True, use_container_width=True)
        else:
            st.success("All required slots assigned ✓")

    with st.expander(f"Avoid-rotation assignments ({len(avoid_assigns)})", expanded=False):
        if avoid_assigns:
            df = pd.DataFrame(
                [
                    {"date": d.isoformat() if hasattr(d, "isoformat") else d,
                     "resident": resident, "rotation": rotation, "slot": slot}
                    for d, resident, rotation, slot in avoid_assigns
                ]
            )
            st.dataframe(df, hide_index=True, use_container_width=True)
        else:
            st.caption("No AVOID-rotation assignments.")

    with st.expander("Run metadata", expanded=False):
        meta = {
            "seed": audit.get("seed"),
            "tiebreaker_count": audit.get("tiebreaker_count"),
            "swap_improvements": audit.get("swap_improvements"),
            "monte_carlo_score_order": audit.get("monte_carlo_score_order"),
            "pick_candidate_rank_order": audit.get("pick_candidate_rank_order"),
            "pick_candidate_weights": audit.get("pick_candidate_weights"),
            "intern_block1_weekday_calls": audit.get("intern_block1_weekday_calls"),
            "block1_end": audit.get("block1_end"),
            "restart_date": audit.get("restart_date"),
            "completed_call_count": audit.get("completed_call_count"),
        }
        st.json(meta)


def _totals_dataframe(residents: dict) -> pd.DataFrame:
    rows = []
    for name, r in residents.items():
        rows.append({
            "name": name,
            "pgy": r.get("pgy"),
            "total_calls": r.get("total_calls", 0),
            "weekday_calls": r.get("weekday_calls", 0),
            "weekend_calls": r.get("weekend_calls", 0),
            "friday_calls": r.get("friday_calls", 0),
            "saturday_calls": r.get("saturday_calls", 0),
            "sunday_calls": r.get("sunday_calls", 0),
            "Jul_Dec_calls": r.get("Jul_Dec_calls", 0),
            "Jan_Jun_calls": r.get("Jan_Jun_calls", 0),
        })
    return pd.DataFrame(rows)


def _render_totals_table(result: dict) -> None:
    df = _totals_dataframe(result["residents"])
    if df.empty:
        st.caption("No residents loaded.")
        return

    # Totals row (sum of numeric columns) appended at the bottom.
    sums = df.select_dtypes(include="number").sum(numeric_only=True)
    totals_row = {col: ("" if col in ("name", "pgy") else int(sums.get(col, 0))) for col in df.columns}
    totals_row["name"] = "TOTAL"
    df_with_totals = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)

    def _row_style(row):
        pgy = row.get("pgy")
        color = _PGY_ROW_COLORS.get(pgy)
        if row["name"] == "TOTAL":
            return [
                "background-color: #EEEEEE; font-weight: bold;" for _ in row
            ]
        if color:
            return [f"background-color: {color};" for _ in row]
        return ["" for _ in row]

    styler = df_with_totals.style.apply(_row_style, axis=1)
    st.dataframe(styler, hide_index=True, use_container_width=True)


def _schedule_dataframe(result: dict) -> pd.DataFrame:
    """Build the schedule view. Mirrors call_schedule.xlsx columns.

    One row per calendar day with Block / Date / Day / Upper level /
    Intern / No Call columns, matching how the user sees the schedule
    in Excel.
    """
    schedule_rows = result["schedule_rows"]
    lookup = result["lookup"]
    no_call = result["no_call"]
    holidays = result["holidays"]
    residents = result["residents"]

    intern_names = [n for n, r in residents.items() if r.get("pgy") == 1]
    nf_name = str(st.session_state.config.get("NIGHT_FLOAT_ROTATION_NAME", "NF"))

    by_date: dict[date, dict] = {}
    completed_dates: set[date] = set()
    for r in schedule_rows:
        d = date.fromisoformat(r["date"])
        slot = r["slot"]
        name = (r.get("resident") or "").strip()
        rec = by_date.setdefault(d, {"upper": "", "intern_weekend": "", "intern_weekday": ""})
        if slot in ("UPPER_WEEKDAY", "UPPER_WEEKEND"):
            rec["upper"] = name
        elif slot == "INTERN_WEEKEND":
            rec["intern_weekend"] = name
        elif slot == "INTERN_WEEKDAY":
            rec["intern_weekday"] = name
        if r.get("note") == "COMPLETED":
            completed_dates.add(d)

    date_to_block: dict[date, int] = {}
    for i, block in enumerate(lookup.blocks, start=1):
        cur = block.start
        while cur <= block.end:
            date_to_block[cur] = i
            cur += timedelta(days=1)

    rows = []
    for d in sorted(by_date.keys()):
        upper = by_date[d]["upper"]
        if d.weekday() >= 5:
            intern_val = by_date[d]["intern_weekend"] or "0"
        else:
            intern_weekday = by_date[d].get("intern_weekday", "")
            if intern_weekday:
                intern_val = intern_weekday
            else:
                nf_interns = [
                    name for name in intern_names
                    if lookup.rotation_on_date(name, d) == nf_name
                ]
                intern_val = ", ".join(sorted(nf_interns)) if nf_interns else "0"

        no_call_entries = []
        for name, days in no_call.items():
            if d in days:
                reason = days[d]
                no_call_entries.append(f"{name} ({reason})" if reason else name)

        if d in holidays:
            note = "HOLIDAY"
        elif d in completed_dates:
            note = "COMPLETED"
        elif not upper or (d.weekday() >= 5 and not by_date[d]["intern_weekend"]):
            note = "UNASSIGNED"
        else:
            note = ""

        rows.append({
            "Block": date_to_block.get(d, ""),
            "Date": d.isoformat(),
            "Day": d.strftime("%a"),
            "Upper level": upper,
            "Intern": intern_val,
            "No Call": ", ".join(sorted(no_call_entries)),
            "_month": d.strftime("%B"),
            "_weekend": d.weekday() >= 5,
            "_note": note,
        })
    return pd.DataFrame(rows)


def _render_schedule_table(result: dict) -> None:
    df = _schedule_dataframe(result)
    if df.empty:
        st.caption("No schedule rows.")
        return

    months_present = list(dict.fromkeys(df["_month"].tolist()))
    residents_present = sorted(
        {x for x in df["Upper level"].tolist() if x and x != "0"}
        | {x for x in df["Intern"].tolist() if x and x not in ("0",)}
    )

    cols = st.columns([1, 1, 1, 1])
    month = cols[0].selectbox("Month", ["All"] + months_present, key="sched_month")
    resident = cols[1].selectbox("Resident", ["All"] + residents_present, key="sched_resident")
    notes_only = cols[2].checkbox("Notes only (HOLIDAY/COMPLETED/UNASSIGNED)", key="sched_notes_only")
    cols[3].caption(f"{len(df)} total rows")

    view = df.copy()
    if month != "All":
        view = view[view["_month"] == month]
    if resident != "All":
        view = view[(view["Upper level"] == resident) | (view["Intern"].str.contains(resident, na=False))]
    if notes_only:
        view = view[view["_note"] != ""]

    visible_cols = ["Block", "Date", "Day", "Upper level", "Intern", "No Call"]
    display = view[visible_cols + ["_weekend", "_note"]].reset_index(drop=True)

    def _row_style(row):
        note = row["_note"]
        if note == "UNASSIGNED":
            return ["background-color: #F4CCCC;" for _ in row]
        if note == "HOLIDAY":
            return ["background-color: #CFE2F3;" for _ in row]
        if note == "COMPLETED":
            return ["background-color: #EBEBEB;" for _ in row]
        if row["_weekend"]:
            return ["background-color: #FFF2CC;" for _ in row]
        return ["" for _ in row]

    styler = display.style.apply(_row_style, axis=1).hide(axis="columns", subset=["_weekend", "_note"])
    st.dataframe(styler, hide_index=True, use_container_width=True, height=560)


def _render_results_section() -> None:
    state = st.session_state.get("run_state", "idle")
    result = st.session_state.get("last_result")
    if state == "running":
        st.info("Generating new schedule…")
        return
    if result is None:
        st.caption("Run a schedule above to see results here.")
        return

    _render_status_banner(result)
    _render_downloads()

    with st.expander("Call totals", expanded=True):
        _render_totals_table(result)
    with st.expander("Call schedule", expanded=False):
        _render_schedule_table(result)
    with st.expander("Audit", expanded=False):
        _render_audit_view(result)


# ---------------------------------------------------------------------------
# Page
# ---------------------------------------------------------------------------


def main() -> None:
    st.set_page_config(page_title="Call Schedule Creator", layout="wide")
    _init_session_state()
    # Drain progress/log/done events from the background thread BEFORE
    # rendering any section, so the Run section sees the freshest state.
    _drain_queue()

    header_cols = st.columns([5, 1, 1])
    header_cols[0].title("Call Schedule Creator")
    if header_cols[1].button("Reset session"):
        _reset_session()
        st.rerun()
    # Exit button: explicit shutdown so non-technical users don't have
    # the asymmetric "close browser, terminal still alive in the background"
    # experience. Requires a confirm step because os._exit kills the process
    # immediately — if a run is in flight, the user would lose it.
    if header_cols[2].button("Exit", type="secondary"):
        st.session_state["_confirm_exit"] = True
        st.rerun()

    if st.session_state.get("_confirm_exit"):
        st.warning(
            "Click **Confirm exit** to shut down the app. The browser tab "
            "will stop responding once the server stops — close it manually. "
            "Any unsaved settings or in-progress runs will be lost."
        )
        confirm_cols = st.columns([1, 1, 6])
        if confirm_cols[0].button("Confirm exit", type="primary"):
            # os._exit bypasses cleanup hooks — that's intentional. Streamlit's
            # tornado server doesn't expose a graceful shutdown API from inside
            # a script, and sys.exit/SystemExit gets caught by Streamlit's own
            # exception handler (the script just reruns). os._exit is the only
            # reliable way to terminate the python.exe spawned by run.bat.
            import os as _os
            _os._exit(0)
        if confirm_cols[1].button("Cancel"):
            st.session_state["_confirm_exit"] = False
            st.rerun()

    with st.expander("1. Upload input files", expanded=False):
        for slot in UPLOAD_SLOTS:
            _render_upload_slot(slot)

    with st.expander("2. Settings", expanded=False):
        settings_has_errors, soft_warnings = _render_settings_section()

    with st.expander("3. Run schedule", expanded=False):
        uploads_ready = _all_required_valid()
        _render_run_section(uploads_ready, settings_has_errors, soft_warnings)

    with st.expander("4. Results", expanded=False):
        _render_results_section()

    # Live-progress tick: while a run is in flight, sleep briefly and
    # rerun so the next pass picks up new queue events. Decision in
    # handoff option (a) — no extra dep, keeps the page responsive
    # enough at ~2.5 reruns/sec.
    if st.session_state.get("run_state") == "running":
        time.sleep(RUN_POLL_INTERVAL_SEC)
        st.rerun()


if __name__ == "__main__":
    main()
