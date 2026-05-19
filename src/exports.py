from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import CONFIG

NIGHT_FLOAT_ROTATION_NAME = CONFIG.get("NIGHT_FLOAT_ROTATION_NAME", "NF")
DATA_DIR = CONFIG.get("DATA_DIR", "data")
OUTPUT_DIR = CONFIG.get("OUTPUT_DIR", "output")


def _autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def _style_header(ws):
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_call_totals_xlsx(residents: dict, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = [
        "name",
        "pgy",
        "total_calls",
        "weekday_calls",
        "weekend_calls",
        "friday_calls",
        "saturday_calls",
        "Jul_Dec_calls",
        "Jan_Jun_calls",
    ]

    ws.append(headers)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    pgy1_fill = PatternFill("solid", fgColor="D9EAD3")
    pgy2_fill = PatternFill("solid", fgColor="CFE2F3")
    pgy3_fill = PatternFill("solid", fgColor="F4CCCC")

    # Vertical separators on the right edge of pgy (B), total_calls (C),
    # weekend_calls (E), and saturday_calls (G) — groups:
    # identity | total | weekday/weekend | friday/saturday | halves.
    divider_cols = (2, 3, 5, 7)
    right_border = Border(right=Side(style="thin"))

    for name, r in residents.items():
        ws.append([
            name,
            r["pgy"],
            r["total_calls"],
            r["weekday_calls"],
            r["weekend_calls"],
            r.get("friday_calls", 0),
            r.get("saturday_calls", 0),
            r["Jul_Dec_calls"],
            r["Jan_Jun_calls"],
        ])

        row_i = ws.max_row
        if r["pgy"] == 1:
            fill = pgy1_fill
        elif r["pgy"] == 2:
            fill = pgy2_fill
        else:
            fill = pgy3_fill

        for col in range(1, 10):
            ws.cell(row=row_i, column=col).fill = fill

        for col in divider_cols:
            ws.cell(row=row_i, column=col).border = right_border

    # Apply the same dividers to the header row so the lines are continuous.
    for col in divider_cols:
        ws.cell(row=1, column=col).border = right_border

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 20

    wb.save(path)


def write_call_schedule_xlsx(
    schedule_rows: list[dict],
    holidays: dict,
    no_call_days: dict,
    path: str,
    lookup,
    intern_names: list[str],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Call List"

    headers = ["Block", "Date", "Day", "Upper level", "Intern", "No Call"]
    ws.append(headers)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    by_date = {}
    completed_dates: set = set()
    for r in schedule_rows:
        d = date.fromisoformat(r["date"])
        slot = r["slot"]
        name = (r.get("resident") or "").strip()

        rec = by_date.setdefault(d, {"upper": "", "intern_weekend": "", "intern_weekday": ""})
        if slot in ("UPPER_WEEKDAY", "UPPER_WEEKEND"):
            rec["upper"] = name
        elif slot == "INTERN_WEEKEND":
            rec["intern_weekend"] = name
        elif slot == "INTERN_WEEKDAY":
            rec["intern_weekday"] = name

        if r.get("note") == "COMPLETED":
            completed_dates.add(d)

    weekend_fill = PatternFill("solid", fgColor="FFF2CC")
    holiday_fill = PatternFill("solid", fgColor="F8CBAD")
    block_fill = PatternFill("solid", fgColor="D9EAD3")
    completed_fill = PatternFill("solid", fgColor="EBEBEB")  # light grey for completed rows
    thick_top = Border(top=Side(style="medium"))

    date_to_block_num = {}
    for i, block in enumerate(lookup.blocks, start=1):
        cur = block.start
        while cur <= block.end:
            date_to_block_num[cur] = i
            cur += timedelta(days=1)

    block_start_dates = {block.start for block in lookup.blocks}

    row_i = 2
    for d in sorted(by_date.keys()):
        upper = by_date[d]["upper"]

        if d.weekday() >= 5:
            intern_val = by_date[d]["intern_weekend"] or "0"
        else:
            intern_weekday = by_date[d].get("intern_weekday", "")
            if intern_weekday:
                # Block 1 weekday intern call — show the assigned resident
                intern_val = intern_weekday
            else:
                nf_interns = []
                for name in intern_names:
                    rot = lookup.rotation_on_date(name, d)
                    if rot == NIGHT_FLOAT_ROTATION_NAME:
                        nf_interns.append(name)
                intern_val = ", ".join(sorted(nf_interns)) if nf_interns else "0"

        no_call_entries = []
        for name, days in no_call_days.items():
            if d in days:
                reason = days[d]
                if reason:
                    no_call_entries.append(f"{name} ({reason})")
                else:
                    no_call_entries.append(name)

        no_call_val = ", ".join(sorted(no_call_entries))
        day_name = d.strftime("%a")
        block_num = date_to_block_num.get(d, "")
        ws.append([block_num, d.isoformat(), day_name, upper, intern_val, no_call_val])

        if d in holidays:
            fill = holiday_fill
        elif d in completed_dates:
            fill = completed_fill
        elif d.weekday() >= 5:
            fill = weekend_fill
        else:
            fill = None

        if fill:
            for col in range(1, 7):
                ws.cell(row=row_i, column=col).fill = fill

        # Italicise name cells on completed rows so they read as "locked"
        if d in completed_dates:
            italic = Font(italic=True, color="808080")
            for col in (4, 5):  # Upper level, Intern
                ws.cell(row=row_i, column=col).font = italic

        if d in block_start_dates:
            for col in range(1, 6):
                ws.cell(row=row_i, column=col).border = thick_top
                if col == 1:
                    ws.cell(row=row_i, column=col).fill = block_fill
                    ws.cell(row=row_i, column=col).font = Font(bold=True)

        row_i += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 40

    wb.save(path)


def write_audit(audit_data, path=f"{DATA_DIR}/{OUTPUT_DIR}/audit_report.txt"):
    with open(path, "w", encoding="utf-8") as f:
        f.write("SCHEDULE AUDIT REPORT\n")
        f.write("=" * 60 + "\n\n")

        f.write("SCHEDULE GENERATION INFO\n")
        f.write("-" * 60 + "\n")
        f.write(f"Seed used: {audit_data['seed']}\n")
        f.write(f"Tie-break decisions: {audit_data['tiebreaker_count']}\n\n")

        f.write("PICK CANDIDATE CONFIGURATION\n")
        f.write("-" * 60 + "\n")
        pick_rank_order = audit_data.get("pick_candidate_rank_order", [])
        pick_weights = audit_data.get("pick_candidate_weights", {})

        f.write(f"Pick candidate rank order: {pick_rank_order}\n")
        for key, value in pick_weights.items():
            f.write(f"{key}: {value}\n")
        f.write("\n")

        f.write("MONTE CARLO CONFIGURATION\n")
        f.write("-" * 60 + "\n")
        score_order = audit_data.get("monte_carlo_score_order", [])
        f.write(f"Monte Carlo score order: {score_order}\n\n")

        f.write("FLOW SHEET INFO\n")
        f.write("-" * 60 + "\n")
        skipped_rows = audit_data.get("skipped_rows", [])
        if skipped_rows:
            f.write(f"Skipped Excel rows: {', '.join(str(r) for r in skipped_rows)}\n")
        else:
            f.write("Skipped Excel rows: None\n")
        f.write("\n")

        f.write("FAIRNESS SUMMARY\n")
        f.write("-" * 60 + "\n")
        for key, value in audit_data["fairness_summary"].items():
            f.write(f"{key}: {value}\n")

        f.write("\nUNASSIGNED SLOTS\n")
        f.write("-" * 60 + "\n")
        if audit_data["unassigned_rows"]:
            for row in audit_data["unassigned_rows"]:
                f.write(
                    f"{row.get('date', '')} | "
                    f"{row.get('slot', '')} | "
                    f"{row.get('holiday', '')} | "
                    f"{row.get('reasons', '')}\n"
                )
        else:
            f.write("No unassigned slots.\n")

        f.write("\nAVOID ASSIGNMENTS USED\n")
        f.write("-" * 60 + "\n")
        if audit_data["avoid_assignments"]:
            for d, resident, rotation, slot in audit_data["avoid_assignments"]:
                f.write(f"{d} | {slot} | {resident} | {rotation}\n")
        else:
            f.write("No AVOID assignments used.\n")

        f.write("\nWEEKEND CALLS > 4 IN A MONTH\n")
        f.write("-" * 60 + "\n")
        if audit_data.get("weekend_call_overages"):
            for row in audit_data["weekend_call_overages"]:
                f.write(
                    f"{row['resident']} | {row['month']} | "
                    f"{row['weekend_calls']} weekend call shifts\n"
                )
        else:
            f.write("No residents exceeded 4 weekend call shifts in any month.\n")

        f.write("\nROTATION DATE SUMMARY\n")
        f.write("-" * 60 + "\n")
        for resident, segments in audit_data.get("rotation_date_summary", {}).items():
            f.write(f"{resident}\n")
            for seg in segments:
                if seg["parts_total"] == 1:
                    label = f"  Block {seg['block']}"
                else:
                    label = f"  Block {seg['block']} part {seg['part']}/{seg['parts_total']}"
                f.write(
                    f"{label}: {seg['start']} -> {seg['end']} | {seg['rotation']}\n"
                )
            f.write("\n")

        f.write("ERRORS\n")
        f.write("-" * 60 + "\n")
        if audit_data["errors"]:
            for e in audit_data["errors"]:
                f.write(f"{e}\n")
        else:
            f.write("No hard-rule violations found.\n")

        f.write("\nWARNINGS\n")
        f.write("-" * 60 + "\n")
        if audit_data["warnings"]:
            for w in audit_data["warnings"]:
                f.write(f"{w}\n")
        else:
            f.write("No warnings.\n")

    print("\nAUDIT COMPLETE")
    print(f"Errors: {len(audit_data['errors'])}")
    print(f"Warnings: {len(audit_data['warnings'])}")
    skipped_rows = audit_data.get("skipped_rows", [])
    if skipped_rows:
        print(f"Skipped Excel rows: {', '.join(str(r) for r in skipped_rows)}")
    else:
        print("Skipped Excel rows: None")
    print(f"Upper total diff: {audit_data['fairness_summary']['upper_total_diff']}")
    print(f"Upper weekday diff: {audit_data['fairness_summary']['upper_weekday_diff']}")
    print(f"Upper weekend diff: {audit_data['fairness_summary']['upper_weekend_diff']}")
    print(f"Intern weekend diff: {audit_data['fairness_summary']['intern_weekend_diff']}")
    print(f"Unassigned slots: {len(audit_data['unassigned_rows'])}")
    print(f"AVOID assignments used: {len(audit_data['avoid_assignments'])}")
    print(f"Weekend-call monthly overages: {len(audit_data.get('weekend_call_overages', []))}")
    print(f"Tie-break decisions: {audit_data['tiebreaker_count']}")
    print(f"Pick candidate rank order: {audit_data.get('pick_candidate_rank_order', [])}")
    print(f"Pick candidate weights: {audit_data.get('pick_candidate_weights', {})}")
    print(f"Monte Carlo score order: {audit_data.get('monte_carlo_score_order', [])}")
    print(f"Audit report written to: {path}")
