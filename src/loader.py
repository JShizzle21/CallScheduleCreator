import re
import warnings
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple, Iterable
from openpyxl import load_workbook

from config import CONFIG
from errors import DataValidationError

ROTATION_RULES_XLSX = CONFIG.get("ROTATION_RULES_XLSX", "input_files/rotation_rules.xlsx")
NO_CALL_DAYS_XLSX = CONFIG.get("NO_CALL_DAYS_XLSX", "input_files/no_call_days.xlsx")
HOLIDAYS_XLSX = CONFIG.get("HOLIDAYS_XLSX", "input_files/holidays.xlsx")
CLINIC_DAYS_XLSX = CONFIG.get("CLINIC_DAYS_XLSX", "input_files/clinic_days.xlsx")
COMPLETED_CALLS_XLSX = CONFIG.get("COMPLETED_CALLS_XLSX", "")


def load_residents(lookup) -> Dict[str, dict]:
    ws = lookup.ws
    skip_rows = lookup.skip_rows
    separator_pgy_labels: dict = getattr(lookup, "separator_pgy_labels", {})

    residents = {}
    pgy = 1
    separator_count = 0

    for r in range(3, ws.max_row + 1):
        if r in skip_rows:
            separator_count += 1
            explicit_pgy = separator_pgy_labels.get(r)
            if explicit_pgy is not None:
                pgy = explicit_pgy
            else:
                pgy += 1
                if pgy > 3:
                    raise ValueError(
                        f"Flow sheet has {separator_count} PGY separator rows, implying "
                        f"PGY{pgy} which is unexpected. Expected at most 2 separators "
                        f"(for PGY1/PGY2/PGY3). Check for extra header rows, or add "
                        f"explicit PGY labels (e.g. 'PGY2') in column A of separator rows."
                    )
            continue

        name = ws.cell(row=r, column=1).value
        if not name:
            continue

        name = str(name).strip()

        residents[name] = {
            "pgy": pgy,
            "assigned_dates": [],
            "total_calls": 0,
            "weekday_calls": 0,
            "weekend_calls": 0,
            "friday_calls": 0,
            "saturday_calls": 0,
            "sunday_calls": 0,
            "upper_calls": 0,
            "intern_calls": 0,
            "Jul_Dec_calls": 0,
            "Jan_Jun_calls": 0,
        }

    return residents


def _normalize_header(value) -> str:
    return str(value).strip().lower() if value is not None else ""


_DATE_FORMATS = (
    "%Y-%m-%d",        # 2026-07-01
    "%Y/%m/%d",        # 2026/07/01
    "%m/%d/%Y",        # 7/1/2026
    "%m/%d/%y",        # 7/1/26
    "%m-%d-%Y",        # 7-1-2026
    "%m-%d-%y",        # 7-1-26
    "%B %d %Y",        # July 1 2026
    "%B %d, %Y",       # July 1, 2026
    "%d %B %Y",        # 1 July 2026
    "%d-%B-%Y",        # 1-July-2026
    "%b %d %Y",        # Jul 1 2026
    "%b %d, %Y",       # Jul 1, 2026
    "%d %b %Y",        # 1 Jul 2026
    "%d-%b-%Y",        # 1-Jul-2026
)


def _parse_date(value) -> date:
    if value is None or value == "":
        raise ValueError("Blank date cell encountered while loading Excel input file.")

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    s = str(value).strip()
    # Collapse internal whitespace so "July  1,  2026" still parses.
    s_clean = " ".join(s.split())
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s_clean, fmt).date()
        except ValueError:
            pass

    raise ValueError(f"Could not parse date value '{value}' from Excel input file.")


def _iter_excel_dict_rows(path: str) -> Iterable[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = [_normalize_header(cell.value) for cell in ws[1]]
    if not any(headers):
        return

    for r in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue
        yield {headers[i]: row_values[i] for i in range(len(headers))}


def load_rotation_rules(path: str = ROTATION_RULES_XLSX) -> Dict[Tuple[str, int], str]:
    rules: Dict[Tuple[str, int], str] = {}
    for row in _iter_excel_dict_rows(path):
        rotation = str(row["rotation_name"]).strip()
        pgy = int(row["pgy"])
        pref = str(row["preference"]).strip().upper()
        rules[(rotation, pgy)] = pref
    return rules


_NO_CALL_HEADER_SCAN_ROWS = 30
_NO_CALL_REQUIRED_HEADERS = ("first name", "last name", "start date", "end date")
_NO_CALL_DEFAULT_TYPE = "time off"


def _find_no_call_header_row(ws, sheet_name: str, search_after: int = 0) -> Optional[int]:
    """Locate a no-call header row (First Name / Last Name / Start Date / End Date).

    Scans up to _NO_CALL_HEADER_SCAN_ROWS rows starting after `search_after`,
    looking for a row that contains ALL four required header tokens
    (case-insensitive, whitespace-collapsed). Returns the row number, or
    None if not found.
    """
    max_row = min(ws.max_row, search_after + _NO_CALL_HEADER_SCAN_ROWS)
    for r in range(search_after + 1, max_row + 1):
        cells = [
            " ".join(str(ws.cell(r, c).value or "").lower().split())
            for c in range(1, ws.max_column + 1)
        ]
        if all(token in cells for token in _NO_CALL_REQUIRED_HEADERS):
            return r
    return None


def _no_call_column_map(ws, header_row: int) -> Dict[str, int]:
    """Map of required-header → 1-based column index."""
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        token = " ".join(str(ws.cell(header_row, c).value or "").lower().split())
        if token in _NO_CALL_REQUIRED_HEADERS:
            out.setdefault(token, c)
    return out


def load_no_call_days(
    path: str = NO_CALL_DAYS_XLSX,
    *,
    valid_residents: Optional[Iterable[str]] = None,
    academic_start: Optional[date] = None,
    academic_end: Optional[date] = None,
    blocks: Optional[List] = None,
) -> Tuple[Dict[str, Dict[date, str]], List[dict]]:
    """Load no-call (time-off) requests from the 13-sheet workbook.

    Workbook format:
      - Sheets named 'Block 1' through 'Block 13' (case-insensitive).
      - Per sheet, two sections — "Interns" and "Uppers" — each with its
        own header row containing columns First Name, Last Name, Start
        Date, End Date. Section labels are decoration only and may sit
        anywhere; the loader scans for the header row.
      - Optional `type` column to the right of End Date (free-form);
        defaults to 'time off' when absent.

    Name matching:
      Only the Last Name column is matched against `valid_residents`.
      Match is case-insensitive and whitespace-trimmed. Unknown last
      names raise DataValidationError with a sheet/row pointer.

    Validations (when params are supplied):
      - Both dates must fall inside the academic year → hard error.
      - A range crossing the Block N sheet's calendar range emits a
        warning but is accepted as written (per-day rows are still
        blocked across the full range).

    Returns ``(per_day, entries)``:
      - ``per_day``: ``{canonical_resident: {date: type_str}}`` — the
        existing API consumed by the scheduler/audit.
      - ``entries``: list of dicts with keys ``resident``, ``start``,
        ``end``, ``type``, ``sheet``, ``row`` — preserved range entries
        for the audit report.

    Empty sheets and missing block sheets are tolerated (warns).
    """
    per_day: Dict[str, Dict[date, str]] = {}
    entries: List[dict] = []

    try:
        wb = load_workbook(path, data_only=True)
    except FileNotFoundError:
        return per_day, entries

    # Case-insensitive last-name lookup → canonical (flow-sheet) spelling.
    last_name_lookup: Optional[Dict[str, str]] = None
    if valid_residents is not None:
        last_name_lookup = {
            n.strip().lower(): n for n in valid_residents
        }

    sheetnames_lower = {s.lower(): s for s in wb.sheetnames}

    for block_idx in range(1, EXPECTED_CLINIC_BLOCK_COUNT + 1):
        expected = f"Block {block_idx}"
        actual_name = sheetnames_lower.get(expected.lower())
        if actual_name is None:
            raise DataValidationError(
                f"no_call_days.xlsx is missing sheet '{expected}'. "
                f"The workbook must contain 13 sheets named 'Block 1' "
                f"through 'Block 13'."
            )

        ws = wb[actual_name]
        block = blocks[block_idx - 1] if blocks is not None and len(blocks) >= block_idx else None

        # Find every header row in the sheet (one per section). We don't
        # care about the section labels themselves — just keep walking
        # downward looking for header rows.
        header_rows: List[int] = []
        cursor = 0
        while True:
            hdr = _find_no_call_header_row(ws, actual_name, search_after=cursor)
            if hdr is None:
                break
            header_rows.append(hdr)
            cursor = hdr

        if not header_rows:
            # Empty sheet (no headers found) — treat as zero entries.
            continue

        # For each header row, scan downward until the next header row
        # (or end of sheet). Blank rows are skipped.
        for section_idx, hdr_row in enumerate(header_rows):
            colmap = _no_call_column_map(ws, hdr_row)
            if len(colmap) < len(_NO_CALL_REQUIRED_HEADERS):
                # Should not happen because _find_no_call_header_row
                # already required all four — defensive check.
                continue

            first_col = colmap["first name"]
            last_col = colmap["last name"]
            start_col = colmap["start date"]
            end_col = colmap["end date"]

            # Optional 'type' column — anywhere to the right of end date
            type_col: Optional[int] = None
            for c in range(1, ws.max_column + 1):
                token = " ".join(str(ws.cell(hdr_row, c).value or "").lower().split())
                if token == "type":
                    type_col = c
                    break

            section_end = (
                header_rows[section_idx + 1] - 1
                if section_idx + 1 < len(header_rows)
                else ws.max_row
            )

            for r in range(hdr_row + 1, section_end + 1):
                first_val = ws.cell(r, first_col).value
                last_val = ws.cell(r, last_col).value
                start_val = ws.cell(r, start_col).value
                end_val = ws.cell(r, end_col).value

                # Row entirely blank? Skip silently (working-doc spacing).
                if all(v is None or str(v).strip() == "" for v in
                       (first_val, last_val, start_val, end_val)):
                    continue

                # Partial row? Must have a last name + both dates.
                last_raw = str(last_val or "").strip()
                if not last_raw:
                    # First-name-only or stray row — skip silently.
                    if not (start_val or end_val):
                        continue
                    raise DataValidationError(
                        f"no_call_days.xlsx sheet '{actual_name}' row {r}: "
                        f"last name is blank but date values are present."
                    )

                if start_val is None or end_val is None or \
                        str(start_val).strip() == "" or str(end_val).strip() == "":
                    raise DataValidationError(
                        f"no_call_days.xlsx sheet '{actual_name}' row {r}: "
                        f"name '{last_raw}' is missing a start or end date."
                    )

                try:
                    start = _parse_date(start_val)
                    end = _parse_date(end_val)
                except ValueError as e:
                    raise DataValidationError(
                        f"no_call_days.xlsx sheet '{actual_name}' row {r}: "
                        f"could not parse date — {e}"
                    ) from e

                if end < start:
                    raise DataValidationError(
                        f"no_call_days.xlsx sheet '{actual_name}' row {r}: "
                        f"end date {end} is before start date {start}."
                    )

                if academic_start is not None and academic_end is not None:
                    if start < academic_start or end > academic_end:
                        raise DataValidationError(
                            f"no_call_days.xlsx sheet '{actual_name}' row {r}: "
                            f"range {start} -> {end} is outside the academic "
                            f"year ({academic_start} to {academic_end})."
                        )

                # Last-name canonicalisation against the flow sheet.
                if last_name_lookup is not None:
                    canonical = last_name_lookup.get(last_raw.lower())
                    if canonical is None:
                        raise DataValidationError(
                            f"no_call_days.xlsx sheet '{actual_name}' row {r}: "
                            f"resident last name '{last_raw}' does not match "
                            f"any resident in the flow sheet. Names must "
                            f"match (case-insensitive)."
                        )
                    resident = canonical
                else:
                    resident = last_raw

                reason_raw = ""
                if type_col is not None:
                    reason_raw = str(ws.cell(r, type_col).value or "").strip()
                reason = reason_raw or _NO_CALL_DEFAULT_TYPE

                # Cross-block warning (range straddles the block this
                # sheet represents).
                if block is not None:
                    if start < block.start or end > block.end:
                        warnings.warn(
                            f"no_call_days.xlsx sheet '{actual_name}' "
                            f"row {r}: {resident} range {start} -> {end} "
                            f"crosses this block's calendar range "
                            f"({block.start} to {block.end}). "
                            f"Accepted as written.",
                            stacklevel=2,
                        )

                # Fan out into per-day map.
                cur = start
                while cur <= end:
                    per_day.setdefault(resident, {})[cur] = reason
                    cur += timedelta(days=1)

                entries.append({
                    "resident": resident,
                    "first_name": str(first_val or "").strip(),
                    "start": start,
                    "end": end,
                    "type": reason,
                    "sheet": actual_name,
                    "row": r,
                })

    return per_day, entries


_HOLIDAY_HEADER_LABELS = {"holiday", "holidays"}
_PGY_SECTION_LABELS = {"pgy-1", "pgy-2", "pgy-3", "pgy1", "pgy2", "pgy3"}
_TOTAL_COLUMN_KEYWORDS = ("total",)
_ER_PREFIX_PATTERN = re.compile(r"^er(\s|$)", re.IGNORECASE)


def _is_er_rotation(value: str) -> bool:
    """True if a rotation cell counts as an ER call assignment.

    Matches case-insensitively any cell that starts with the token 'ER'
    followed by either end-of-string or whitespace. So 'ER', 'er',
    'ER 24', 'er pm' all match. 'ERIC' or 'ERAS' do NOT match.
    """
    return bool(_ER_PREFIX_PATTERN.match(value.strip()))


def _parse_holiday_date_with_year_inference(
    value, academic_start: Optional[date], academic_end: Optional[date]
) -> Optional[date]:
    """Parse a holiday date cell, tolerating decorative wrappers.

    The supervisor sometimes wraps draft dates in asterisks or other
    punctuation to flag them as "tentative" (e.g. ``****June 4****``).
    This helper:
      1. Tries the normal _parse_date first.
      2. Strips common punctuation wrappers and retries.
      3. If the resulting string has no year (e.g. ``June 4``), inserts
         each candidate year from the academic window in turn and
         accepts whichever lands inside the window.

    Returns the parsed date, or None if all attempts failed.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    # First pass: try the value as-is.
    try:
        return _parse_date(value)
    except ValueError:
        pass

    # Strip non-alphanumeric/space wrappers and retry as a year-bearing string.
    cleaned = re.sub(r"[^A-Za-z0-9/\-\s,]", " ", str(value)).strip()
    cleaned = " ".join(cleaned.split())
    if not cleaned:
        return None

    try:
        return _parse_date(cleaned)
    except ValueError:
        pass

    # No year in the string — try inferring from the academic window.
    if academic_start is not None and academic_end is not None:
        for year in (academic_start.year, academic_end.year):
            candidate = f"{cleaned} {year}"
            try:
                parsed = _parse_date(candidate)
            except ValueError:
                continue
            if academic_start <= parsed <= academic_end:
                return parsed

    return None


def _find_holiday_header_row(ws, max_scan_rows: int = 15) -> Optional[int]:
    """Locate the row whose column A holds 'Holiday' / 'Holidays'.

    Scans the first `max_scan_rows` rows. Case-insensitive, trimmed.
    Returns 1-based row index or None.
    """
    upper_bound = min(ws.max_row, max_scan_rows)
    for r in range(1, upper_bound + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        if str(v).strip().lower() in _HOLIDAY_HEADER_LABELS:
            return r
    return None


def _find_holiday_date_row(
    ws, header_row: int, holiday_columns: List[int],
    academic_start: Optional[date], academic_end: Optional[date],
) -> Optional[int]:
    """Find the row containing parseable dates beneath the Holiday header.

    Per spec the date row is 1 or 2 rows below the Holiday header. The
    helper scans both candidate rows and picks the one where at least one
    of the holiday columns yields a parseable date.
    """
    for offset in (1, 2):
        candidate = header_row + offset
        if candidate > ws.max_row:
            return None
        for col in holiday_columns:
            v = ws.cell(row=candidate, column=col).value
            if _parse_holiday_date_with_year_inference(
                v, academic_start, academic_end
            ) is not None:
                return candidate
    return None


def load_holidays(
    path: str = HOLIDAYS_XLSX,
    *,
    valid_residents: Optional[Iterable[str]] = None,
    intern_names: Optional[Iterable[str]] = None,
    academic_start: Optional[date] = None,
    academic_end: Optional[date] = None,
    known_rotations: Optional[Iterable[str]] = None,
) -> Tuple[Dict[date, dict], Dict[str, Dict[date, str]]]:
    """Load the holiday-rotation workbook → (holidays_dict, pre_holiday_no_call).

    Workbook layout (one sheet, working-doc style):
      - A row in column A labelled 'Holiday' (or 'Holidays', case-insensitive)
        within the first ~15 rows. Columns to the right of A on that row hold
        the holiday display names.
      - 1 or 2 rows below the Holiday header row: a row of dates (one per
        holiday column). Stray cells like ``****June 4****`` are tolerated
        (asterisks stripped, year inferred from the academic window).
      - Rows below the date row: PGY section markers ('PGY-1', 'PGY-2',
        'PGY-3') interleaved with resident rows. Resident rows have a last
        name in column A; each holiday column holds the rotation that
        resident is working on the holiday (blank = not working).
      - A right-most informational column commonly headed 'Total Holidays'
        is detected and skipped.

    Validation:
      - `valid_residents`: every resident in this set must appear in the
        workbook (hard error if missing); every name in the workbook must
        match this set (hard error if unknown). Last-name comparison is
        case-insensitive + whitespace-trimmed.
      - `known_rotations`: rotation cell values not in this set emit a
        warning (NOT a hard error). ER-prefixed values are exempt.

    ER handling:
      - Any cell whose value starts with 'ER' (case-insensitive, optional
        space-anything suffix) marks the resident as the ER call for that
        holiday. ER residents are sorted into ``uppers`` (PGY2/3) or
        ``interns`` (PGY1) lists on the holiday's dict entry. Multiple ER
        residents in the same PGY band on the same holiday are accepted
        (the audit will surface the count mismatch but execution
        continues).

    Day-before block:
      - For every resident with a non-empty rotation cell on holiday D
        (ER or not), date D-1 is added to a per-resident no-call map
        with reason 'pre_holiday'. The scheduler merges this into its
        eligibility filter so the resident isn't scheduled for regular
        call the night before they work a holiday rotation.

    Returns:
      - ``holidays_dict[d] = {
            "name": str,
            "uppers": [pgy23_er_names],
            "interns": [pgy1_er_names],
            "rotations": {resident: rotation_str},
        }``
      - ``pre_holiday_no_call[resident] = {date: 'pre_holiday'}``
    """
    holidays_out: Dict[date, dict] = {}
    pre_holiday_no_call: Dict[str, Dict[date, str]] = {}

    try:
        wb = load_workbook(path, data_only=True)
    except FileNotFoundError:
        return holidays_out, pre_holiday_no_call

    ws = wb.active

    header_row = _find_holiday_header_row(ws)
    if header_row is None:
        raise DataValidationError(
            "holidays.xlsx: could not find a 'Holiday' header in column A "
            "within the first 15 rows. The header row tells the loader "
            "where to find holiday names and dates."
        )

    # Identify holiday columns from the header row. A holiday column is
    # any column (B+) whose holiday-name cell is non-empty and doesn't
    # look like a total column.
    holiday_columns: List[int] = []
    holiday_names: Dict[int, str] = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None or str(v).strip() == "":
            continue
        name_str = str(v).strip()
        if any(kw in name_str.lower() for kw in _TOTAL_COLUMN_KEYWORDS):
            continue
        holiday_columns.append(c)
        holiday_names[c] = name_str

    if not holiday_columns:
        raise DataValidationError(
            f"holidays.xlsx: 'Holiday' header found at row {header_row} "
            f"but no holiday-name columns to the right. Expected one "
            f"column per holiday with the holiday name in row {header_row}."
        )

    date_row = _find_holiday_date_row(
        ws, header_row, holiday_columns, academic_start, academic_end
    )
    if date_row is None:
        raise DataValidationError(
            f"holidays.xlsx: 'Holiday' header at row {header_row} but no "
            f"parseable date row found in the next two rows. The date row "
            f"should contain one date per holiday column."
        )

    # Resolve each holiday column → its date (or skip with warning).
    column_dates: Dict[int, date] = {}
    for col in holiday_columns:
        raw = ws.cell(row=date_row, column=col).value
        parsed = _parse_holiday_date_with_year_inference(
            raw, academic_start, academic_end
        )
        if parsed is None:
            warnings.warn(
                f"holidays.xlsx column {col} ('{holiday_names[col]}'): "
                f"cell at row {date_row} is not a parseable date "
                f"(value={raw!r}). Skipping this holiday.",
                stacklevel=2,
            )
            continue
        if academic_start is not None and academic_end is not None:
            if parsed < academic_start or parsed > academic_end:
                warnings.warn(
                    f"holidays.xlsx column {col} ('{holiday_names[col]}'): "
                    f"date {parsed} is outside the academic year "
                    f"({academic_start} to {academic_end}). Skipping.",
                    stacklevel=2,
                )
                continue
        column_dates[col] = parsed

    # Last-name canonicalisation map: lowercase → canonical (flow) spelling.
    last_name_lookup: Optional[Dict[str, str]] = None
    flow_names_canonical: set = set()
    if valid_residents is not None:
        last_name_lookup = {n.strip().lower(): n for n in valid_residents}
        flow_names_canonical = set(valid_residents)

    intern_set = set(intern_names) if intern_names is not None else None

    rotation_lookup_lower: Optional[set] = None
    if known_rotations is not None:
        rotation_lookup_lower = {r.strip().lower() for r in known_rotations if r}

    seen_residents: set = set()
    unknown_rotations_warned: set = set()

    # Process rows below the date row. Skip PGY section markers; expect
    # resident-name + per-holiday rotation cells everywhere else.
    for r in range(date_row + 1, ws.max_row + 1):
        col_a = ws.cell(row=r, column=1).value
        if col_a is None or str(col_a).strip() == "":
            continue
        col_a_str = str(col_a).strip()
        lowered = col_a_str.lower().replace(" ", "")
        if lowered in _PGY_SECTION_LABELS:
            continue

        # Treat col A as a resident last name.
        if last_name_lookup is not None:
            canonical = last_name_lookup.get(col_a_str.lower())
            if canonical is None:
                raise DataValidationError(
                    f"holidays.xlsx row {r}: resident last name "
                    f"'{col_a_str}' does not match any resident in the "
                    f"flow sheet. Names must match (case-insensitive)."
                )
            resident = canonical
        else:
            resident = col_a_str

        seen_residents.add(resident)

        # Process each holiday column for this resident.
        for col, d in column_dates.items():
            v = ws.cell(row=r, column=col).value
            if v is None or str(v).strip() == "":
                continue
            rotation_str = str(v).strip()

            # Initialise the holiday's entry if needed.
            entry = holidays_out.setdefault(d, {
                "name": holiday_names[col],
                "uppers": [],
                "interns": [],
                "rotations": {},
            })

            # Same date, different sheet column: shouldn't happen but if
            # two columns resolve to the same date, the first name wins.
            entry["rotations"][resident] = rotation_str

            is_er = _is_er_rotation(rotation_str)

            # Day-before block applies to every listed resident, regardless
            # of ER status.
            d_minus_1 = d - timedelta(days=1)
            if academic_start is None or d_minus_1 >= academic_start:
                pre_holiday_no_call.setdefault(resident, {})[d_minus_1] = "pre_holiday"

            if is_er:
                if intern_set is not None and resident in intern_set:
                    entry["interns"].append(resident)
                else:
                    entry["uppers"].append(resident)
            else:
                # Non-ER rotation — emit a warning if it doesn't look like
                # any known rotation code from flow / rotation_rules.
                if rotation_lookup_lower is not None:
                    if rotation_str.lower() not in rotation_lookup_lower:
                        key = (resident, d, rotation_str)
                        if key not in unknown_rotations_warned:
                            unknown_rotations_warned.add(key)
                            warnings.warn(
                                f"holidays.xlsx row {r} col {col} "
                                f"({holiday_names[col]} {d}): rotation "
                                f"'{rotation_str}' for {resident} doesn't "
                                f"match any known rotation code. Accepted "
                                f"as written.",
                                stacklevel=2,
                            )

    # Cross-check: every resident in flow must be listed.
    if last_name_lookup is not None:
        missing = sorted(flow_names_canonical - seen_residents)
        if missing:
            raise DataValidationError(
                f"holidays.xlsx is missing rows for {len(missing)} "
                f"resident(s) from the flow sheet: "
                f"{', '.join(missing)}. Every resident must have a row "
                f"(blank cells are OK; the row just needs to exist)."
            )

    return holidays_out, pre_holiday_no_call


CLINIC_HEADER_SCAN_ROWS = 10
EXPECTED_CLINIC_BLOCK_COUNT = 13

# Body-row cell values that match these patterns are silently ignored — they
# are header-like labels (or stray copy-paste artefacts) from the
# supervisor's working doc, not resident names.
_CLINIC_IGNORED_BODY_LABELS = {"date", "intern", "interns"}


# Module-level set of clinic-warning keys already emitted in this process.
# Prevents the same '?' warning from firing once per Monte Carlo seed when
# run_simulation rebuilds the data bundle per run.
# (Each ProcessPoolExecutor worker has its own copy, so the warning will
# still fire once per worker process — acceptably small.)
_clinic_warning_seen: set = set()

# Also register Python's own per-process "once" filter as a belt-and-braces
# guard in case our explicit set is somehow bypassed.
warnings.filterwarnings(
    "once",
    message=r".*uncertain entry.*Confirm with supervisor\.",
    category=UserWarning,
)


def _is_ignored_clinic_label(value: str) -> bool:
    """True if a body-row cell value is a known non-name label to skip."""
    lower = value.lower()
    if lower in _CLINIC_IGNORED_BODY_LABELS:
        return True
    # 'Amb - <whatever>' is a rotation/group marker the supervisor places in
    # the sheet — never a resident name on its own.
    if lower.startswith("amb -") or lower.startswith("amb-"):
        return True
    return False


def _is_parseable_date(value) -> bool:
    """Lightweight helper: True iff value parses as a date via _parse_date."""
    if value is None or value == "":
        return False
    try:
        _parse_date(value)
        return True
    except ValueError:
        return False


def _find_clinic_date_column(ws, sheet_name: str) -> Optional[Tuple[int, int]]:
    """Locate the 'date' header in a clinic-days block sheet.

    Scans the first CLINIC_HEADER_SCAN_ROWS rows for any cell whose
    case-insensitive trimmed value equals 'date'. Each candidate is
    classified by the cell directly below:
      - Parses as a date            → CONFIRMED (the right column to use).
      - Blank/empty                 → AMBIGUOUS (could be an empty sheet).
      - Non-empty, not a date       → BAD (suspicious — header is misplaced
                                       or the column is mislabelled).

    Returns:
      - (header_row, date_col) for the first confirmed candidate.
      - None if no 'date' header exists at all OR all 'date' headers are
        AMBIGUOUS (treated as "this block has no clinics yet").

    Raises DataValidationError only when a 'date' header has non-date data
    below it AND no other column confirms — that pattern usually means the
    user has the date column in the wrong place.
    """
    confirmed: Optional[Tuple[int, int]] = None
    bad: List[Tuple[int, int, object]] = []
    ambiguous_count = 0

    max_scan_row = min(CLINIC_HEADER_SCAN_ROWS, ws.max_row)
    for r in range(1, max_scan_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip().lower() != "date":
                continue
            below = ws.cell(row=r + 1, column=c).value
            if below is None or str(below).strip() == "":
                ambiguous_count += 1
                continue
            if _is_parseable_date(below):
                if confirmed is None:
                    confirmed = (r, c)
            else:
                bad.append((r, c, below))

    if confirmed is not None:
        return confirmed

    if bad:
        details = ", ".join(
            f"row {r} col {c} (below = {value!r})" for r, c, value in bad
        )
        raise DataValidationError(
            f"clinic_days.xlsx sheet '{sheet_name}': found 'Date' header(s) "
            f"at {details}, but the cell below is not a recognizable date. "
            f"Either fix the value or move the 'Date' header to the correct "
            f"column."
        )

    # No confirmed and no bad. Either there were only ambiguous 'date'
    # headers (empty block), or no 'date' header at all (also treated as
    # empty — the loader will still error if the sheet has stray names
    # elsewhere).
    return None


def load_clinic_days(
    path: str = CLINIC_DAYS_XLSX,
    *,
    valid_residents: Optional[Iterable[str]] = None,
    academic_start: Optional[date] = None,
    academic_end: Optional[date] = None,
    blocks: Optional[List] = None,
) -> Dict[str, Dict[date, str]]:
    """Load clinic days from a 13-sheet workbook → day-before no-call map.

    Residency rule: a resident cannot take call the night before a clinic
    because the following morning they must be present in clinic. The day
    before each clinic date is blocked with reason 'pre_clinic_day'.

    Workbook format:
      - Sheets named 'Block 1' through 'Block 13' (one per academic block).
      - Each sheet has a 'Date' column (case-insensitive header) somewhere in
        the first 10 rows; the column may be preceded by other columns (e.g.
        day-of-week) that are ignored.
      - The cell directly below the 'Date' header confirms the column —
        it must parse as a date.
      - Rows below the header: column = a clinic date; cells to the right
        list the resident names with clinic that day. Empty cells are
        skipped. Cells to the left of the Date column are ignored.
      - If a row's date cell is blank or unparseable, the row is skipped
        silently (allows in-progress edits).

    Validation (only when the corresponding params are supplied):
      - `valid_residents`: any name not in this set → DataValidationError.
        Also caps the rightward scan at `len(valid_residents)` cells past
        the Date column.
      - `academic_start`/`academic_end`: clinic dates outside this inclusive
        range → DataValidationError.
      - `blocks`: list of Block objects (with .start/.end). A date on sheet
        'Block N' that falls outside blocks[N-1]'s range → DataValidationError.

    If the file does not exist, returns an empty dict (clinics are optional).
    """
    out: Dict[str, Dict[date, str]] = {}
    try:
        wb = load_workbook(path, data_only=True)
    except FileNotFoundError:
        return out

    name_set = set(valid_residents) if valid_residents is not None else None
    resident_count = len(name_set) if name_set is not None else None

    sheetnames_lower = {s.lower(): s for s in wb.sheetnames}

    for block_idx in range(1, EXPECTED_CLINIC_BLOCK_COUNT + 1):
        expected = f"Block {block_idx}"
        actual_name = sheetnames_lower.get(expected.lower())
        if actual_name is None:
            raise DataValidationError(
                f"clinic_days.xlsx is missing sheet '{expected}'. "
                f"The workbook must contain 13 sheets named 'Block 1' "
                f"through 'Block 13'."
            )

        ws = wb[actual_name]
        located = _find_clinic_date_column(ws, actual_name)
        if located is None:
            # No confirmed date column. Treat as empty block UNLESS the sheet
            # has stray non-header content past the scan rows, which would
            # suggest data is present but the Date column is missing.
            scan_floor = CLINIC_HEADER_SCAN_ROWS + 1
            has_stray = False
            for r in range(scan_floor, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is not None and str(v).strip() != "":
                        has_stray = True
                        break
                if has_stray:
                    break
            if has_stray:
                raise DataValidationError(
                    f"clinic_days.xlsx sheet '{actual_name}': contains data "
                    f"below row {CLINIC_HEADER_SCAN_ROWS} but no 'Date' "
                    f"header with a recognizable date value could be found "
                    f"in the first {CLINIC_HEADER_SCAN_ROWS} rows. Add a "
                    f"'Date' column header with a valid date below it."
                )
            continue
        header_row, date_col = located

        # Restrict rightward scan: never more cells than total residents.
        if resident_count is not None:
            scan_end = min(ws.max_column, date_col + resident_count)
        else:
            scan_end = ws.max_column

        block = blocks[block_idx - 1] if blocks is not None and len(blocks) >= block_idx else None

        for r in range(header_row + 1, ws.max_row + 1):
            date_cell = ws.cell(row=r, column=date_col).value
            if date_cell is None or str(date_cell).strip() == "":
                continue
            try:
                clinic_date = _parse_date(date_cell)
            except ValueError:
                # Non-date value in a body row — treat as in-progress / ignore.
                continue

            if academic_start is not None and academic_end is not None:
                if clinic_date < academic_start or clinic_date > academic_end:
                    raise DataValidationError(
                        f"clinic_days.xlsx sheet '{actual_name}' row {r}: "
                        f"clinic date {clinic_date} is outside the academic "
                        f"year ({academic_start} to {academic_end})."
                    )

            if block is not None:
                if clinic_date < block.start or clinic_date > block.end:
                    raise DataValidationError(
                        f"clinic_days.xlsx sheet '{actual_name}' row {r}: "
                        f"clinic date {clinic_date} is outside this block's "
                        f"calendar range ({block.start} to {block.end}). "
                        f"Move this row to the correct block sheet."
                    )

            blocked_date = clinic_date - timedelta(days=1)
            for c in range(date_col + 1, scan_end + 1):
                v = ws.cell(row=r, column=c).value
                if v is None or str(v).strip() == "":
                    continue
                raw = str(v).strip()

                # Skip header-like labels that bleed into body cells
                # (working-doc artefacts: stray 'Date', 'Intern',
                # 'Amb - <something>', etc.).
                if _is_ignored_clinic_label(raw):
                    continue

                # Trailing '?' marks an uncertain entry from the supervisor's
                # working doc. Strip the '?' for matching purposes (the
                # source file is left untouched), still apply the call, and
                # emit a one-time warning so the user can confirm later.
                # The '?' stays in the supervisor's spreadsheet; only our
                # in-memory copy is stripped.
                name = raw
                if name.endswith("?"):
                    name = name.rstrip("?").strip()
                    key = (actual_name, r, raw)
                    if key not in _clinic_warning_seen:
                        _clinic_warning_seen.add(key)
                        warnings.warn(
                            f"clinic_days.xlsx sheet '{actual_name}' row {r}: "
                            f"uncertain entry '{raw}' - treating as '{name}'. "
                            f"Confirm with supervisor.",
                            stacklevel=2,
                        )

                if name_set is not None and name not in name_set:
                    raise DataValidationError(
                        f"clinic_days.xlsx sheet '{actual_name}' row {r}: "
                        f"resident name '{name}' does not match any resident "
                        f"in the flow sheet. Names must match exactly."
                    )
                out.setdefault(name, {})[blocked_date] = "pre_clinic_day"

    return out


def load_completed_calls(path: str = COMPLETED_CALLS_XLSX, block1_end: Optional[date] = None) -> List[Tuple[date, str, str]]:
    """Load a partially-completed call schedule to seed a mid-year restart.

    File format: one row per call day, with columns:
      - a 'date' column (any header containing 'date')
      - an upper-level column (any header containing 'upper')
      - an intern column (any header containing 'intern')
    Blank cells mean no assignment for that slot on that date.

    Slot types are inferred from the date: weekday → UPPER_WEEKDAY,
    weekend upper → UPPER_WEEKEND, weekend intern → INTERN_WEEKEND.

    Returns a list of (date, resident_name, slot) tuples sorted by date.
    Returns an empty list if the file does not exist or has no data.
    """
    if not path:
        return []

    out: List[Tuple[date, str, str]] = []
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        headers = [_normalize_header(cell.value) for cell in ws[1]]
        if not any(headers):
            return []

        date_col: Optional[int] = next(
            (i for i, h in enumerate(headers) if "date" in h), None
        )
        upper_col: Optional[int] = next(
            (i for i, h in enumerate(headers) if "upper" in h), None
        )
        intern_col: Optional[int] = next(
            (i for i, h in enumerate(headers) if "intern" in h), None
        )

        if date_col is None:
            raise ValueError(
                f"No column containing 'date' found in {path}. "
                f"Headers found: {headers}"
            )

        for r in range(2, ws.max_row + 1):
            row_vals = [
                ws.cell(row=r, column=c + 1).value for c in range(len(headers))
            ]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue

            try:
                d = _parse_date(row_vals[date_col])
            except (ValueError, TypeError):
                continue

            if upper_col is not None:
                upper_name = row_vals[upper_col]
                if upper_name and str(upper_name).strip():
                    slot = "UPPER_WEEKEND" if d.weekday() >= 5 else "UPPER_WEEKDAY"
                    out.append((d, str(upper_name).strip(), slot))

            if intern_col is not None:
                intern_name = row_vals[intern_col]
                if intern_name and str(intern_name).strip():
                    if d.weekday() >= 5:
                        out.append((d, str(intern_name).strip(), "INTERN_WEEKEND"))
                    elif block1_end is not None and d <= block1_end:
                        # Block 1 intern weekday call (feature enabled)
                        out.append((d, str(intern_name).strip(), "INTERN_WEEKDAY"))
                    # else: weekday intern = Night Float — silently dropped

    except FileNotFoundError:
        pass

    return sorted(out, key=lambda x: x[0])
