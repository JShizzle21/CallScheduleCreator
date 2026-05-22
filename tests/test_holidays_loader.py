"""Tests for the holidays-rotation workbook loader and the day-before
holiday + NF no-call blocks.

The loader has a lot of moving parts — header scanning, date-row scanning,
ER-prefix detection, asterisk-wrapped dates, last-name canonicalisation,
hard-error on missing/unknown residents, rotation-name warnings — so the
unit tests construct workbooks programmatically to exercise each branch.

The integration tests then run the full pipeline against the real input
files in data/ and verify the day-of/day-before holiday rules and the
pre-NF intern rule are actually applied in the produced schedule.
"""

from __future__ import annotations

from datetime import date, datetime

import pytest
from openpyxl import Workbook

from errors import DataValidationError
from loader import load_holidays, _is_er_rotation


def _build_holidays_workbook(
    tmp_path,
    holiday_columns,        # list of (name, date)
    resident_rows,          # list of (last_name, pgy_label, {hol_idx: rotation_str})
    *,
    decorative_header_rows=0,  # how many blank rows to put above 'Holiday'
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_row = decorative_header_rows + 1
    ws.cell(row=header_row, column=1).value = "Holiday"
    for i, (name, _d) in enumerate(holiday_columns):
        ws.cell(row=header_row, column=2 + i).value = name
    # Add a fake "Total Holidays" trailing column to confirm it's skipped.
    ws.cell(row=header_row, column=2 + len(holiday_columns)).value = "Total Holidays"

    # Date row goes 2 rows below the header (with a decorative day-of-week
    # row sandwiched in between, mirroring the supervisor's file).
    dow_row = header_row + 1
    date_row = header_row + 2
    for i, (_name, d) in enumerate(holiday_columns):
        ws.cell(row=dow_row, column=2 + i).value = d.strftime("%A")
        ws.cell(row=date_row, column=2 + i).value = datetime.combine(
            d, datetime.min.time()
        )

    # Resident rows, grouped by PGY label.
    cur_row = date_row + 1
    last_pgy = None
    for last_name, pgy, rotations in resident_rows:
        if pgy != last_pgy:
            ws.cell(row=cur_row, column=1).value = pgy
            cur_row += 1
            last_pgy = pgy
        ws.cell(row=cur_row, column=1).value = last_name
        for hol_idx, rot in rotations.items():
            ws.cell(row=cur_row, column=2 + hol_idx).value = rot
        cur_row += 1

    path = tmp_path / "holidays.xlsx"
    wb.save(path)
    return path


# -----------------------------------------------------------------------
# Unit: ER detection
# -----------------------------------------------------------------------

@pytest.mark.parametrize("value,expected", [
    ("ER", True),
    ("er", True),
    ("ER 24", True),
    ("er 24", True),
    ("ER PM", True),
    ("er pm", True),
    (" ER ", True),
    ("", False),
    ("OB", False),
    ("INP", False),
    ("ERIC", False),       # 'ER' must be followed by space or end
    ("ERAS", False),
    ("HER", False),
])
def test_er_rotation_detection(value, expected):
    assert _is_er_rotation(value) is expected


# -----------------------------------------------------------------------
# Unit: loader produces correct ER assignments + day-before blocks
# -----------------------------------------------------------------------

def test_er_residents_become_uppers_or_interns(tmp_path):
    path = _build_holidays_workbook(
        tmp_path,
        holiday_columns=[
            ("Independence Day", date(2026, 7, 4)),
            ("Labor Day", date(2026, 9, 7)),
        ],
        resident_rows=[
            ("Storey",  "PGY-1", {0: "ER 24"}),
            ("Haque",   "PGY-1", {1: "er pm"}),
            ("McMurray", "PGY-3", {0: "ER 24"}),
            ("Green",   "PGY-2", {1: "ER 24"}),
        ],
    )

    valid = {"Storey", "Haque", "McMurray", "Green"}
    intern_names = {"Storey", "Haque"}
    holidays, pre_holiday = load_holidays(
        str(path),
        valid_residents=valid,
        intern_names=intern_names,
        academic_start=date(2026, 7, 1),
        academic_end=date(2027, 6, 30),
    )

    indep = holidays[date(2026, 7, 4)]
    assert indep["name"] == "Independence Day"
    assert indep["interns"] == ["Storey"]
    assert indep["uppers"] == ["McMurray"]
    assert indep["rotations"] == {"Storey": "ER 24", "McMurray": "ER 24"}

    labor = holidays[date(2026, 9, 7)]
    assert labor["interns"] == ["Haque"]
    assert labor["uppers"] == ["Green"]

    # Day-before blocks applied to every listed resident.
    assert date(2026, 7, 3) in pre_holiday["Storey"]
    assert date(2026, 7, 3) in pre_holiday["McMurray"]
    assert pre_holiday["Storey"][date(2026, 7, 3)] == "pre_holiday"
    assert date(2026, 9, 6) in pre_holiday["Haque"]


def test_non_er_residents_are_blocked_but_not_assigned(tmp_path):
    path = _build_holidays_workbook(
        tmp_path,
        holiday_columns=[("Christmas Day", date(2026, 12, 25))],
        resident_rows=[
            ("Loera", "PGY-1", {0: "INP"}),  # blocked, not ER
            ("Haque", "PGY-1", {0: "ER 24"}),  # ER → intern call
        ],
    )

    holidays, pre_holiday = load_holidays(
        str(path),
        valid_residents={"Loera", "Haque"},
        intern_names={"Loera", "Haque"},
        academic_start=date(2026, 7, 1),
        academic_end=date(2027, 6, 30),
    )

    christmas = holidays[date(2026, 12, 25)]
    assert christmas["interns"] == ["Haque"]
    assert "Loera" not in christmas["interns"]
    # But Loera IS day-before-blocked.
    assert date(2026, 12, 24) in pre_holiday["Loera"]
    assert date(2026, 12, 24) in pre_holiday["Haque"]


# -----------------------------------------------------------------------
# Unit: name validation
# -----------------------------------------------------------------------

def test_unknown_resident_hard_errors(tmp_path):
    path = _build_holidays_workbook(
        tmp_path,
        holiday_columns=[("Independence Day", date(2026, 7, 4))],
        resident_rows=[
            ("Storey", "PGY-1", {0: "ER 24"}),
            ("Nobody", "PGY-1", {0: "INP"}),
        ],
    )

    with pytest.raises(DataValidationError, match="Nobody"):
        load_holidays(
            str(path),
            valid_residents={"Storey"},  # Nobody is not in flow
            intern_names={"Storey"},
            academic_start=date(2026, 7, 1),
            academic_end=date(2027, 6, 30),
        )


def test_missing_resident_hard_errors(tmp_path):
    """Every resident from the flow sheet must appear in the file."""
    path = _build_holidays_workbook(
        tmp_path,
        holiday_columns=[("Independence Day", date(2026, 7, 4))],
        resident_rows=[
            ("Storey", "PGY-1", {0: "ER 24"}),
        ],
    )

    with pytest.raises(DataValidationError, match="missing rows"):
        load_holidays(
            str(path),
            valid_residents={"Storey", "Haque", "McMurray"},
            intern_names={"Storey", "Haque"},
            academic_start=date(2026, 7, 1),
            academic_end=date(2027, 6, 30),
        )


def test_last_name_match_is_case_insensitive(tmp_path):
    """Working-doc casing/whitespace tolerance for last names."""
    path = _build_holidays_workbook(
        tmp_path,
        holiday_columns=[("Independence Day", date(2026, 7, 4))],
        resident_rows=[
            ("storey ", "PGY-1", {0: "ER 24"}),  # trailing space + lowercase
        ],
    )

    holidays, _ = load_holidays(
        str(path),
        valid_residents={"Storey"},  # canonical
        intern_names={"Storey"},
        academic_start=date(2026, 7, 1),
        academic_end=date(2027, 6, 30),
    )
    # Resolved to the canonical flow-sheet spelling.
    assert holidays[date(2026, 7, 4)]["interns"] == ["Storey"]


# -----------------------------------------------------------------------
# Unit: asterisk-wrapped date with year inference
# -----------------------------------------------------------------------

def test_asterisk_wrapped_date_parses_with_year_inference(tmp_path):
    """``****June 4****`` should be accepted as June 4 of the year that
    lands inside the academic window — academic year 2026-07 .. 2027-06
    means June 4 = 2027-06-04."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "Holiday"
    ws.cell(row=1, column=2).value = "Graduation"
    ws.cell(row=2, column=2).value = "Friday"   # day-of-week row
    ws.cell(row=3, column=2).value = "****June 4****"
    ws.cell(row=4, column=1).value = "PGY-1"
    ws.cell(row=5, column=1).value = "Storey"
    ws.cell(row=5, column=2).value = "ER 24"
    path = tmp_path / "holidays.xlsx"
    wb.save(path)

    holidays, _ = load_holidays(
        str(path),
        valid_residents={"Storey"},
        intern_names={"Storey"},
        academic_start=date(2026, 7, 1),
        academic_end=date(2027, 6, 30),
    )
    assert date(2027, 6, 4) in holidays
    assert holidays[date(2027, 6, 4)]["interns"] == ["Storey"]


# -----------------------------------------------------------------------
# Integration: full pipeline against real input files
# -----------------------------------------------------------------------

@pytest.fixture(scope="module")
def real_pipeline():
    import scheduler_main as sm
    from config import load_default_config

    config, paths = load_default_config()
    sm._apply_config(config)
    return sm.generate_schedule_once(seed=0, paths=paths, config=config)


def test_real_pipeline_assigns_er_residents_on_holidays(real_pipeline):
    """McMurray and Storey were marked ER 24 on Independence Day in the
    real holidays.xlsx — both should appear in the produced schedule
    on 2026-07-04 with HOLIDAY note."""
    rows = real_pipeline["schedule_rows"]
    indep_rows = [r for r in rows if r["date"] == "2026-07-04"]
    holiday_rows = [r for r in indep_rows if (r.get("note") or "").startswith("HOLIDAY")]
    residents = {r["resident"] for r in holiday_rows if r.get("resident")}
    assert "McMurray" in residents
    assert "Storey" in residents


def test_real_pipeline_blocks_day_before_holiday(real_pipeline):
    """No resident listed on a holiday in holidays.xlsx may be assigned
    to regular call the day before that holiday. This is the
    'pre_holiday' no-call block."""
    rows = real_pipeline["schedule_rows"]
    # Day before Independence Day is 2026-07-03. Residents listed on
    # Indep Day include Mahmood, Wallace, Schaeffer, Fuentes, Hall,
    # Prabhu, Storey, McMurray, Godoy.
    blocked = {
        "Mahmood", "Wallace", "Schaeffer", "Fuentes", "Hall",
        "Prabhu", "Storey", "McMurray", "Godoy",
    }
    july3 = [r for r in rows if r["date"] == "2026-07-03"]
    assigned = {r["resident"] for r in july3 if r.get("resident")}
    overlap = blocked & assigned
    assert overlap == set(), (
        f"Residents listed on Independence Day were assigned to call "
        f"the day before: {overlap}"
    )


def test_real_pipeline_blocks_intern_day_before_nf(real_pipeline):
    """Whenever an intern is on NF on day D+1, they shouldn't have a
    regular call assignment on day D. This is the 'pre_nf' no-call
    block."""
    import scheduler_main as sm
    from config import load_default_config
    from datetime import timedelta as _td, date as _date

    # Reuse the bundle's lookup to inspect rotations.
    config, paths = load_default_config()
    from data_bundle import load_data_bundle
    bundle = load_data_bundle(
        paths,
        academic_year_start=int(config["ACADEMIC_DATE_START_STRING"][:4]),
        intern_block1_weekday_calls=bool(int(config.get("INTERN_BLOCK1_WEEKDAY_CALLS", 0))),
        use_completed_calls=False,
        academic_start_date=_date.fromisoformat(config["ACADEMIC_DATE_START_STRING"]),
        academic_end_date=_date.fromisoformat(config["ACADEMIC_DATE_END_STRING"]),
    )
    lookup = bundle.lookup
    interns = [name for name, info in bundle.residents.items() if info.get("pgy") == 1]

    violations = []
    for r in real_pipeline["schedule_rows"]:
        name = (r.get("resident") or "").strip()
        if not name or name not in interns:
            continue
        # Skip pre-applied (HOLIDAY/COMPLETED) rows — those bypass
        # eligibility per user policy.
        note = (r.get("note") or "")
        if note.startswith("HOLIDAY") or note == "COMPLETED":
            continue
        d = _date.fromisoformat(r["date"])
        rot_next = lookup.rotation_on_date(name, d + _td(days=1))
        if rot_next and rot_next.upper() == "NF":
            violations.append((d, name, rot_next))

    assert violations == [], (
        f"Interns assigned to regular call the day before their NF day: "
        f"{violations}"
    )
