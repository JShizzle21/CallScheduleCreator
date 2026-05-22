"""End-to-end tests for the multi-sheet no_call_days.xlsx loader.

Three things are exercised:

1. The loader correctly parses the 13-sheet workbook, finds both
   Interns/Uppers sections via header scanning, canonicalises last
   names (case-insensitive + whitespace), and returns the expected
   per-day map + ranges list.

2. Unknown last names raise a hard error pointing at sheet/row.

3. A full generate_schedule_once run against a real input file
   produces zero "assigned on no-call day" audit errors — i.e. the
   scheduler actually respects the time-off ranges, not just the
   audit picks them up.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from errors import DataValidationError
from loader import load_no_call_days


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"


def _build_workbook(tmp_path: Path, rows_by_block: dict) -> Path:
    """Construct a 13-sheet no-call workbook with `rows_by_block` content.

    `rows_by_block` maps block index (1..13) → list of dicts with keys
    `section` ('Interns'|'Uppers'), `first`, `last`, `start`, `end`,
    `type` (optional).
    """
    wb = Workbook()
    wb.remove(wb.active)
    for i in range(1, 14):
        ws = wb.create_sheet(f"Block {i}")
        ws.cell(row=1, column=1).value = "Interns"
        ws.cell(row=2, column=1).value = "First Name"
        ws.cell(row=2, column=2).value = "Last Name"
        ws.cell(row=2, column=3).value = "Start Date"
        ws.cell(row=2, column=4).value = "End Date"
        ws.cell(row=2, column=5).value = "Type"
        ws.cell(row=10, column=1).value = "Uppers"
        ws.cell(row=11, column=1).value = "First Name"
        ws.cell(row=11, column=2).value = "Last Name"
        ws.cell(row=11, column=3).value = "Start Date"
        ws.cell(row=11, column=4).value = "End Date"
        ws.cell(row=11, column=5).value = "Type"

        intern_r, upper_r = 3, 12
        for entry in rows_by_block.get(i, []):
            target = intern_r if entry["section"] == "Interns" else upper_r
            ws.cell(row=target, column=1).value = entry["first"]
            ws.cell(row=target, column=2).value = entry["last"]
            ws.cell(row=target, column=3).value = datetime.combine(
                entry["start"], datetime.min.time()
            )
            ws.cell(row=target, column=4).value = datetime.combine(
                entry["end"], datetime.min.time()
            )
            if "type" in entry:
                ws.cell(row=target, column=5).value = entry["type"]
            if entry["section"] == "Interns":
                intern_r += 1
            else:
                upper_r += 1

    path = tmp_path / "no_call_days.xlsx"
    wb.save(path)
    return path


def test_loader_parses_multiple_sections_and_canonicalises_names(tmp_path):
    path = _build_workbook(
        tmp_path,
        {
            1: [
                # Intentionally lowercase first + trailing space on last
                {"section": "Interns", "first": "kristelle",
                 "last": "Yelderman ", "start": date(2026, 7, 5),
                 "end": date(2026, 7, 8)},
                {"section": "Uppers", "first": "Cossette",
                 "last": "GREEN", "start": date(2026, 7, 11),
                 "end": date(2026, 7, 13), "type": "vacation"},
            ],
            3: [
                {"section": "Uppers", "first": "Imrana",
                 "last": "Riaz", "start": date(2026, 9, 14),
                 "end": date(2026, 9, 16)},
            ],
        },
    )

    valid = {"Yelderman", "Green", "Riaz"}
    per_day, entries = load_no_call_days(
        str(path),
        valid_residents=valid,
        academic_start=date(2026, 7, 1),
        academic_end=date(2027, 6, 30),
    )

    # All three entries parsed.
    assert len(entries) == 3

    # Last names canonicalised to the flow-sheet spelling.
    assert set(per_day.keys()) == {"Yelderman", "Green", "Riaz"}

    # Each day in each range is present.
    assert per_day["Yelderman"][date(2026, 7, 5)] == "time off"
    assert per_day["Yelderman"][date(2026, 7, 8)] == "time off"
    assert date(2026, 7, 9) not in per_day["Yelderman"]

    # Custom type preserved.
    assert per_day["Green"][date(2026, 7, 11)] == "vacation"

    # Entries list preserves originals with sheet/row pointers.
    green = next(e for e in entries if e["resident"] == "Green")
    assert green["sheet"] == "Block 1"
    assert green["start"] == date(2026, 7, 11)
    assert green["end"] == date(2026, 7, 13)
    assert green["type"] == "vacation"


def test_unknown_last_name_is_hard_error(tmp_path):
    path = _build_workbook(
        tmp_path,
        {
            1: [
                {"section": "Uppers", "first": "Someone",
                 "last": "Nobody", "start": date(2026, 7, 11),
                 "end": date(2026, 7, 13)},
            ],
        },
    )

    with pytest.raises(DataValidationError, match="Nobody"):
        load_no_call_days(
            str(path),
            valid_residents={"Green", "Yelderman"},
            academic_start=date(2026, 7, 1),
            academic_end=date(2027, 6, 30),
        )


def test_out_of_year_date_is_hard_error(tmp_path):
    path = _build_workbook(
        tmp_path,
        {
            1: [
                {"section": "Uppers", "first": "C",
                 "last": "Green", "start": date(2025, 12, 1),
                 "end": date(2025, 12, 5)},
            ],
        },
    )

    with pytest.raises(DataValidationError, match="outside the academic year"):
        load_no_call_days(
            str(path),
            valid_residents={"Green"},
            academic_start=date(2026, 7, 1),
            academic_end=date(2027, 6, 30),
        )


# --------------------------------------------------------------------------
# Full-pipeline integration test: real input files, real scheduler.
# --------------------------------------------------------------------------

@pytest.fixture(scope="module")
def real_schedule():
    """Run the full pipeline once and reuse the result across tests.

    Marked as a module-scoped fixture so the 1000-seed Monte Carlo loop
    only runs once for the whole test module.
    """
    import scheduler_main as sm
    from config import load_default_config

    config, paths = load_default_config()
    sm._apply_config(config)
    return sm.generate_schedule_once(seed=0, paths=paths, config=config)


def test_scheduler_respects_real_time_off_ranges(real_schedule):
    """No resident may be scheduled on any day inside one of their loaded
    no-call ranges.

    The audit emits ``"<date>: <resident> assigned on no-call day"`` for
    any violation. This test fails if any such error appears, which is
    the strongest end-to-end guarantee that the time-off loader and the
    scheduler's eligibility filter agree.
    """
    audit = real_schedule["audit_data"]
    no_call_violations = [
        e for e in audit["errors"] if "assigned on no-call day" in e
    ]
    assert no_call_violations == [], (
        f"Scheduler assigned residents during their time-off ranges:\n"
        + "\n".join(no_call_violations)
    )


def test_audit_lists_time_off_entries(real_schedule):
    """Every range loaded from no_call_days.xlsx is reported back in
    audit_data['no_call_entries'] for sanity-checking against the source."""
    audit = real_schedule["audit_data"]
    entries = audit.get("no_call_entries", [])
    assert entries, "Expected at least one time-off entry in audit output"

    # Each entry has the required keys.
    sample = entries[0]
    for key in ("resident", "start", "end", "type", "sheet", "row"):
        assert key in sample, f"missing key '{key}' in {sample!r}"
