from __future__ import annotations
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook

import re

MONTHS = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
}

#Check if row contains month text indicating it is a header row and not a rotation list row
def looks_like_block_date(text: str) -> bool:
    if not text:
        return False
    text = str(text).strip().upper()
    answer = bool(re.search(r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)", text) and "-" in text)
    return answer


def detect_skip_rows(ws) -> set[int]:
    skip_rows = set()

    for r in range(3, ws.max_row + 1):

        date_like_count = 0
        for c in range(2, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value

            if looks_like_block_date(val):
                date_like_count += 1

        # if several columns look like block dates, treat as separator row
        if date_like_count >= 3:
            skip_rows.add(r)

    return skip_rows

def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())  # weekday: Mon=0 ... Sun=6

def _infer_year(month: int, academic_year_start: int ) -> int:
    # Jul-Dec => start year, Jan-Jun => start year + 1
    return academic_year_start if month >= 7 else academic_year_start + 1

def _to_date(v, academic_year_start: int ) -> date | None:
    """
    Handles:
    - Excel date cells
    - ISO strings like 2026-07-01
    - Date ranges like 'JUL 1 - JUL 27' (returns the start date)
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip().upper()

    # ISO date
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    # Range like 'JUL 1 - JUL 27' (we take the first part)
    # Also tolerates extra spaces and en-dash
    m = re.search(r"\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\b\W*(\d{1,2})", s)
    if m:
        mon_abbr = m.group(1)
        day = int(m.group(2))
        month = MONTHS[mon_abbr]
        year = _infer_year(month, academic_year_start)
        return date(year, month, day)

    return None


@dataclass(frozen=True)
class Block:
    col: int               # Excel column index (1=A)
    start: date
    end: date              # start + 27 days


def load_blocks(xlsx_path: str, sheet_name: str, academic_year_start: int) -> list[Block]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    starts_raw: list[tuple[int, date]] = []
    for col in range(2, ws.max_column + 1):
        d = _to_date(ws.cell(row=2, column=col).value, academic_year_start)
        if d:
            starts_raw.append((col, d))

    # Normalize to Monday
    starts = [(col, monday_of_week(d)) for col, d in starts_raw]
    starts.sort(key=lambda x: x[1])

    final_end = date(academic_year_start + 1, 6, 30)

    blocks: list[Block] = []
    for i, (col, start) in enumerate(starts):
        if i < len(starts) - 1:
            next_start = starts[i + 1][1]          # already Monday
            end = next_start - timedelta(days=1)   # Sunday
        else:
            end = final_end
        blocks.append(Block(col=col, start=start, end=end))

    year_start = date(academic_year_start, 7, 1)
    # after blocks are built:
    blocks[0] = Block(col=blocks[0].col, start=year_start, end=blocks[0].end)

    return blocks



def find_resident_row(ws, skip_rows: set[int], resident_name: str) -> Optional[int]:
    # Names start row 3, but skip relabel rows
    for r in range(3, ws.max_row + 1):
        if r in skip_rows:
            continue
        name = ws.cell(row=r, column=1).value
        if name and str(name).strip() == resident_name:
            return r
    return None


def rotation_on_date(xlsx_path: str, xlsx_sheet_name: str, resident_name: str, skip_rows: set[int], academic_year_start: int, d: date) -> Optional[str]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[xlsx_sheet_name]
    blocks = load_blocks(xlsx_path, xlsx_sheet_name, academic_year_start)

    rr = find_resident_row(ws, skip_rows, resident_name)
    if rr is None:
        return None

    # find which block contains the date
    block = next((b for b in blocks if b.start <= d <= b.end), None)
    if block is None:
        return None

    cell_val = ws.cell(row=rr, column=block.col).value
    if cell_val is None:
        return None

    rot = str(cell_val).strip()
    if "/" in rot:
        first, second = [p.strip() for p in rot.split("/", 1)]
        # first 14 days of block: start..start+13
        if d <= block.start + timedelta(days=13):
            return first
        return second

    return rot


if __name__ == "__main__":
    xlsx = "data/flow.xlsx"  # adjust if needed
    sheet_name = "master_block_calendar"
    resident = "PGY1-5"    # must match column A EXACTLY
    test_date = date(2027, 1, 25)
    rows_skip = {15, 24}  # 1-indexed Excel rows

    wb = load_workbook(xlsx, data_only=True)
    ws = wb["master_block_calendar"]

    # 1) show first few block dates we read from row 2

    blocks = load_blocks(xlsx, sheet_name, 2026)
    print("Blocks found:", len(blocks))
    print("First 13 blocks:", [(b.col, b.start.isoformat(), b.end.isoformat()) for b in blocks[:13]])

    # 2) confirm resident row exists
    rr = find_resident_row(ws, rows_skip, resident)
    print("Resident row:", rr)

    # 3) confirm which block contains test_date
    block = next((b for b in blocks if b.start <= test_date <= b.end), None)
    print("Block for date:", None if block is None else (block.col, block.start, block.end))

    # 4) show the raw cell value for that resident+block
    if rr and block:
        raw = ws.cell(row=rr, column=block.col).value
        print("Raw cell value:", raw)

    # 5) final function result
    print("rotation_on_date:", rotation_on_date(xlsx, sheet_name, resident, rows_skip, 2026, test_date))


@dataclass
class ExcelRotationLookup:
    xlsx_path: str
    sheet_name: str
    academic_year_start: int

    def __post_init__(self):
        # 1) Load once
        self.wb = load_workbook(self.xlsx_path, data_only=True)
        self.ws = self.wb[self.sheet_name]

        self.skip_rows = detect_skip_rows(self.ws)


        # 2) Blocks once
        self.blocks = load_blocks(self.xlsx_path, self.sheet_name, academic_year_start=self.academic_year_start)

        # 3) resident name -> row
        self.resident_row = {}
        for r in range(3, self.ws.max_row + 1):
            if r in self.skip_rows:
                continue
            name = self.ws.cell(row=r, column=1).value
            if name:
                self.resident_row[str(name).strip()] = r


        # 4) date -> (block_col, block_start)
        self.date_to_block = {}
        for b in self.blocks:
            cur = b.start
            while cur <= b.end:
                self.date_to_block[cur] = (b.col, b.start)
                cur += timedelta(days=1)

        # 5) Cache rotation cells: resident -> {block_col: "ROT" }
        self.rotation_grid = {}
        for name, rr in self.resident_row.items():
            per_res = {}
            for b in self.blocks:
                v = self.ws.cell(row=rr, column=b.col).value
                per_res[b.col] = "" if v is None else str(v).strip()
            self.rotation_grid[name] = per_res

        # Precompute: resident -> {date: rotation}
        self.rotation_by_resident_date = {}

        for name in self.resident_row.keys():
            per_day = {}
            for b in self.blocks:
                raw = self.rotation_grid.get(name, {}).get(b.col, "")
                if not raw:
                    continue

                # Determine the rotation for the first and second half of the block
                if "/" in raw:
                    first, second = [p.strip() for p in raw.split("/", 1)]
                else:
                    first, second = raw, raw

                cur = b.start
                while cur <= b.end:
                    # first 14 days use first, then second
                    if cur <= b.start + timedelta(days=13):
                        per_day[cur] = first
                    else:
                        per_day[cur] = second
                    cur += timedelta(days=1)

            self.rotation_by_resident_date[name] = per_day

    def rotation_on_date(self, resident_name: str, d: date) -> Optional[str]:
        return self.rotation_by_resident_date.get(resident_name, {}).get(d)


