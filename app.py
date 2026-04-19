"""Streamlit GUI for the Call Schedule Creator.

Phase 5a — scaffold + upload section only. Settings, Run, and Results are
stubbed. See docs/gui_plan.md for the full spec.

Runtime notes:
- Requires streamlit >= 1.38 (pinned in requirements.txt).
- Uploaded files land in a per-session temp directory created under the
  OS temp location via tempfile.mkdtemp(prefix="CallScheduler_"). On
  Windows this is typically C:\\Users\\<user>\\AppData\\Local\\Temp\\.
  Files in the session tmpdir are deleted when the user clicks
  "Reset session"; orphaned tmpdirs from prior sessions (older than
  24 hours) are cleaned on launch.
- All GUI state lives in st.session_state. Nothing persists to disk
  except via "Save as defaults" (Phase 5b), which overwrites config.yaml.
"""

from __future__ import annotations

import shutil
import tempfile
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable, Optional

import streamlit as st
from openpyxl import load_workbook

from config import load_default_config, save_config

TMPDIR_PREFIX = "CallScheduler_"
ORPHAN_AGE_SECONDS = 24 * 3600

# Fixed allowed sets for the two lexicographic-order fields. The user may
# reorder but not add or remove items — the scheduler requires exactly these.
ALLOWED_MC_SCORE_ORDER = {
    "errors",
    "unassigned",
    "upper_weekend_diff",
    "upper_weekday_diff",
    "pgy2_total_diff",
    "pgy3_total_diff",
    "intern_weekend_diff",
    "avoid_assignments",
    "warnings",
}
ALLOWED_PICK_RANK_ORDER = {
    "hard_diff_flag",
    "soft_diff_flag",
    "weighted_score",
}

# Keys that the Settings section exposes to the user. Path keys and SHEET_NAME
# are intentionally omitted — they're managed by the Upload section.
BEHAVIOR_KEYS = [
    # Common
    "ACADEMIC_DATE_START_STRING",
    "ACADEMIC_DATE_END_STRING",
    "PGY3_CUTOFF_DATE",
    "SIMULATION_RUNS",
    "USE_COMPLETED_CALLS",
    "INTERN_BLOCK1_WEEKDAY_CALLS",
    # Advanced
    "POST_CALL_DAYS",
    "MIN_SPACING_DAYS_STRONG",
    "MIN_SPACING_DAYS_MILD",
    "MAX_CALLS_IN_WINDOW",
    "ROLLING_WINDOW_DAYS",
    "MAX_DIFF_SOFT",
    "MAX_DIFF_HARD",
    "NIGHT_FLOAT_ROTATION_NAME",
    # Expert
    "FAIRNESS_GAP_WEIGHT",
    "SPACING_WEIGHT",
    "AVOID_WEIGHT",
    "YEAR_BIAS_WEIGHT",
    "PACE_WEIGHT",
    "LOOKAHEAD_WEIGHT",
    "PICK_CANDIDATE_RANK_ORDER",
    "MONTE_CARLO_SCORE_ORDER",
]

WEIGHT_KEYS = [
    "FAIRNESS_GAP_WEIGHT",
    "SPACING_WEIGHT",
    "AVOID_WEIGHT",
    "YEAR_BIAS_WEIGHT",
    "PACE_WEIGHT",
    "LOOKAHEAD_WEIGHT",
]


# ---------------------------------------------------------------------------
# Upload slot definitions
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class UploadSlot:
    key: str
    label: str
    saved_filename: str  # What we save the upload as inside the tmpdir.
    required: bool
    validator: Callable[[Path], Optional[str]]
    help: str


def _validate_flow(path: Path) -> Optional[str]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return f"Could not open workbook: {exc}"
    if "master_block_calendar" not in wb.sheetnames:
        return (
            "Flow sheet error: expected a sheet named 'master_block_calendar'. "
            f"Found: {', '.join(wb.sheetnames)}"
        )
    ws = wb["master_block_calendar"]
    # Row 2, column B+ should hold block start dates.
    date_cells = [ws.cell(row=2, column=c).value for c in range(2, min(ws.max_column, 30) + 1)]
    if not any(hasattr(v, "year") for v in date_cells):
        return (
            "Flow sheet error: row 2 does not contain any parseable dates in "
            "columns B onward. Expected block start dates there."
        )
    return None


def _validate_headers(path: Path, required_headers: set[str], label: str) -> Optional[str]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return f"Could not open workbook: {exc}"
    ws = wb.active
    headers = {
        str(ws.cell(row=1, column=c).value).strip().lower()
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value is not None
    }
    missing = {h.lower() for h in required_headers} - headers
    if missing:
        return (
            f"{label} error: header row is missing required column(s): "
            f"{', '.join(sorted(missing))}. Found: {', '.join(sorted(headers))}."
        )
    return None


def _validate_rotation_rules(path: Path) -> Optional[str]:
    return _validate_headers(path, {"rotation", "PGY1", "PGY2", "PGY3"}, "Rotation rules")


def _validate_no_call(path: Path) -> Optional[str]:
    return _validate_headers(path, {"name", "date"}, "No-call days")


def _validate_holidays(path: Path) -> Optional[str]:
    return _validate_headers(path, {"date"}, "Holidays")


def _validate_clinic(path: Path) -> Optional[str]:
    return _validate_headers(path, {"name", "date"}, "Clinic days")


def _validate_completed(path: Path) -> Optional[str]:
    return _validate_headers(path, {"date"}, "Completed calls")


UPLOAD_SLOTS: list[UploadSlot] = [
    UploadSlot(
        key="flow",
        label="Rotation flow sheet",
        saved_filename="flow.xlsx",
        required=True,
        validator=_validate_flow,
        help="Must contain a sheet named 'master_block_calendar' with block start dates in row 2.",
    ),
    UploadSlot(
        key="rotation_rules",
        label="Rotation rules",
        saved_filename="rotation_rules.xlsx",
        required=True,
        validator=_validate_rotation_rules,
        help="Columns: rotation, PGY1, PGY2, PGY3. Values per row: ELIGIBLE, AVOID, or NO_CALL.",
    ),
    UploadSlot(
        key="no_call_days",
        label="No-call days",
        saved_filename="no_call_days.xlsx",
        required=True,
        validator=_validate_no_call,
        help="Columns: name, date. May be empty if no residents have no-call requests.",
    ),
    UploadSlot(
        key="holidays",
        label="Holidays",
        saved_filename="holidays.xlsx",
        required=True,
        validator=_validate_holidays,
        help="Columns: date, name, upper, intern. May be empty if there are no holiday pre-assignments.",
    ),
    UploadSlot(
        key="clinic_days",
        label="Clinic days",
        saved_filename="clinic_days.xlsx",
        required=False,
        validator=_validate_clinic,
        help="Optional. Columns: name, date.",
    ),
    UploadSlot(
        key="completed_calls",
        label="Completed calls",
        saved_filename="completed_calls.xlsx",
        required=False,
        validator=_validate_completed,
        help="Optional. Columns: date, upper, intern. Required only in partial-year mode.",
    ),
]


# ---------------------------------------------------------------------------
# Session-state management
# ---------------------------------------------------------------------------


def _cleanup_orphan_tmpdirs() -> None:
    """Delete CallScheduler_* tmpdirs older than 24 hours.

    Streamlit has no reliable shutdown hook, so prior sessions can leave
    tmpdirs behind. We only touch entries older than ORPHAN_AGE_SECONDS
    so concurrent sessions can't delete each other's state.
    """
    root = Path(tempfile.gettempdir())
    now = time.time()
    for child in root.glob(f"{TMPDIR_PREFIX}*"):
        try:
            if child.is_dir() and now - child.stat().st_mtime > ORPHAN_AGE_SECONDS:
                shutil.rmtree(child, ignore_errors=True)
        except OSError:
            pass


def _init_session_state() -> None:
    if "initialized" in st.session_state:
        return
    _cleanup_orphan_tmpdirs()
    st.session_state.tmpdir = Path(tempfile.mkdtemp(prefix=TMPDIR_PREFIX))
    # uploads[slot.key] = {"saved_path": Path, "original_name": str,
    #                       "uploaded_at": float, "error": Optional[str]}
    st.session_state.uploads = {}
    # Seed config once. Phase 5b will wire widgets to these values.
    config, _paths = load_default_config()
    st.session_state.config = config
    st.session_state.initialized = True


def _reset_session() -> None:
    """Delete the tmpdir and clear session state so the page starts fresh."""
    tmpdir = st.session_state.get("tmpdir")
    if tmpdir is not None:
        shutil.rmtree(tmpdir, ignore_errors=True)
    for key in list(st.session_state.keys()):
        del st.session_state[key]


# ---------------------------------------------------------------------------
# Upload section
# ---------------------------------------------------------------------------


def _save_upload(slot: UploadSlot, uploaded_file) -> None:
    """Persist the Streamlit UploadedFile to the session tmpdir and validate."""
    dest = st.session_state.tmpdir / slot.saved_filename
    with open(dest, "wb") as fh:
        fh.write(uploaded_file.getbuffer())
    error = slot.validator(dest)
    st.session_state.uploads[slot.key] = {
        "saved_path": dest,
        "original_name": uploaded_file.name,
        "uploaded_at": time.time(),
        "error": error,
    }


def _remove_upload(slot_key: str) -> None:
    entry = st.session_state.uploads.pop(slot_key, None)
    if entry is not None:
        try:
            Path(entry["saved_path"]).unlink(missing_ok=True)
        except OSError:
            pass


def _render_upload_slot(slot: UploadSlot) -> None:
    required_tag = "" if slot.required else " _(optional)_"
    st.markdown(f"**{slot.label}**{required_tag}")
    st.caption(slot.help)

    entry = st.session_state.uploads.get(slot.key)
    if entry is not None:
        cols = st.columns([4, 1])
        if entry["error"] is None:
            cols[0].success(
                f"✓ {entry['original_name']} — saved as `{slot.saved_filename}`"
            )
        else:
            cols[0].error(f"✗ {entry['original_name']} — {entry['error']}")
        if cols[1].button("Remove", key=f"remove_{slot.key}"):
            _remove_upload(slot.key)
            st.rerun()
    else:
        st.caption("No file loaded.")

    uploaded = st.file_uploader(
        f"Upload {slot.label}",
        type=["xlsx"],
        key=f"uploader_{slot.key}",
        label_visibility="collapsed",
    )
    if uploaded is not None:
        existing = entry is not None and entry["original_name"] == uploaded.name
        if not existing:
            _save_upload(slot, uploaded)
            st.rerun()

    st.divider()


def _all_required_valid() -> bool:
    for slot in UPLOAD_SLOTS:
        if not slot.required:
            continue
        entry = st.session_state.uploads.get(slot.key)
        if entry is None or entry["error"] is not None:
            return False
    return True


# ---------------------------------------------------------------------------
# Settings — validation
# ---------------------------------------------------------------------------


def _parse_date_or_none(raw) -> Optional[date]:
    """Return a date for a string or date; None if raw is empty/unparseable."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, date):
        return raw
    try:
        return date.fromisoformat(str(raw))
    except ValueError:
        return None


def _is_blank_date(raw) -> bool:
    return raw is None or raw == ""


def _compute_hard_errors(cfg: dict) -> dict[str, list[str]]:
    """Return {key: [messages]} for any hard validation failures.

    Hard errors block the Run button (see §4.6 of docs/gui_plan.md).
    """
    errs: dict[str, list[str]] = defaultdict(list)

    start_raw = cfg.get("ACADEMIC_DATE_START_STRING", "")
    end_raw = cfg.get("ACADEMIC_DATE_END_STRING", "")
    cutoff_raw = cfg.get("PGY3_CUTOFF_DATE", "")

    start = _parse_date_or_none(start_raw)
    end = _parse_date_or_none(end_raw)
    cutoff = _parse_date_or_none(cutoff_raw)

    if start is None:
        errs["ACADEMIC_DATE_START_STRING"].append("Academic year start is not a valid date.")
    if end is None:
        errs["ACADEMIC_DATE_END_STRING"].append("Academic year end is not a valid date.")
    if start and end and end <= start:
        errs["ACADEMIC_DATE_END_STRING"].append(
            "Academic year end must be after academic year start."
        )

    if not _is_blank_date(cutoff_raw):
        if cutoff is None:
            errs["PGY3_CUTOFF_DATE"].append("PGY3 cutoff is not a valid date.")
        elif start and end and not (start <= cutoff <= end):
            errs["PGY3_CUTOFF_DATE"].append(
                "PGY3 cutoff must be between the academic start and end dates "
                "(leave blank to disable)."
            )

    if int(cfg.get("SIMULATION_RUNS", 0)) < 1:
        errs["SIMULATION_RUNS"].append("Simulation runs must be at least 1.")

    for key in (
        "POST_CALL_DAYS",
        "MIN_SPACING_DAYS_STRONG",
        "MIN_SPACING_DAYS_MILD",
        "MAX_CALLS_IN_WINDOW",
        "ROLLING_WINDOW_DAYS",
        "MAX_DIFF_SOFT",
        "MAX_DIFF_HARD",
    ):
        if int(cfg.get(key, 0)) < 0:
            errs[key].append(f"{key} cannot be negative.")

    for key in WEIGHT_KEYS:
        if float(cfg.get(key, 0)) < 0:
            errs[key].append(f"{key} cannot be negative.")

    if int(cfg.get("MAX_DIFF_HARD", 0)) < int(cfg.get("MAX_DIFF_SOFT", 0)):
        errs["MAX_DIFF_HARD"].append("Hard threshold must be ≥ soft threshold.")

    if int(cfg.get("MIN_SPACING_DAYS_MILD", 0)) < int(cfg.get("MIN_SPACING_DAYS_STRONG", 0)):
        errs["MIN_SPACING_DAYS_MILD"].append("Mild spacing must be ≥ strong spacing.")

    if not str(cfg.get("NIGHT_FLOAT_ROTATION_NAME", "")).strip():
        errs["NIGHT_FLOAT_ROTATION_NAME"].append("Night Float rotation name cannot be blank.")

    mc = cfg.get("MONTE_CARLO_SCORE_ORDER")
    if not isinstance(mc, list) or set(mc) != ALLOWED_MC_SCORE_ORDER or len(mc) != len(set(mc)):
        errs["MONTE_CARLO_SCORE_ORDER"].append(
            f"Must list exactly these items once each (any order, one per line): "
            f"{', '.join(sorted(ALLOWED_MC_SCORE_ORDER))}."
        )

    pick = cfg.get("PICK_CANDIDATE_RANK_ORDER")
    if not isinstance(pick, list) or set(pick) != ALLOWED_PICK_RANK_ORDER or len(pick) != len(set(pick)):
        errs["PICK_CANDIDATE_RANK_ORDER"].append(
            f"Must list exactly these items once each (any order, one per line): "
            f"{', '.join(sorted(ALLOWED_PICK_RANK_ORDER))}."
        )

    return dict(errs)


def _compute_soft_warnings(cfg: dict) -> list[str]:
    """Return a list of soft-warning messages (§4.6). Do not block Run."""
    warnings: list[str] = []

    start = _parse_date_or_none(cfg.get("ACADEMIC_DATE_START_STRING"))
    end = _parse_date_or_none(cfg.get("ACADEMIC_DATE_END_STRING"))
    if start and end:
        duration = (end - start).days
        if duration < 300 or duration > 400:
            warnings.append(
                f"Academic year is {duration} days. Typical value is ~365."
            )
        cutoff = _parse_date_or_none(cfg.get("PGY3_CUTOFF_DATE"))
        if cutoff is not None and (end - cutoff).days > 45:
            warnings.append(
                f"PGY3s will be excluded for {(end - cutoff).days} days. "
                "That may strain PGY2 coverage."
            )

    runs = int(cfg.get("SIMULATION_RUNS", 0))
    if runs > 5000:
        warnings.append(
            "Simulation runs > 5000 — expect several minutes. "
            "Consider the Quick preview button if iterating."
        )
    elif 0 < runs < 100:
        warnings.append(
            "Low simulation run count may produce suboptimal schedules. "
            "1000 is the tuned default."
        )

    if int(cfg.get("POST_CALL_DAYS", 0)) == 0:
        warnings.append(
            "Post-call rest days = 0 — residents could be assigned on consecutive days."
        )

    if int(cfg.get("MAX_CALLS_IN_WINDOW", 0)) == 0:
        warnings.append(
            "Rolling-window cap disabled — burst patterns (multiple calls within a short span) "
            "will not be prevented."
        )

    for key in WEIGHT_KEYS:
        val = float(cfg.get(key, 0))
        if val == 0:
            warnings.append(f"{key} is 0 — this component will be ignored during candidate ranking.")
        elif val > 10:
            warnings.append(f"{key} is {val} — far above the default. This component will dominate ranking.")

    return warnings


# ---------------------------------------------------------------------------
# Settings — widget rendering
# ---------------------------------------------------------------------------


def _load_defaults_fresh() -> dict:
    """Re-read config.yaml from disk each rerun for live ● diff comparisons."""
    try:
        cfg, _paths = load_default_config()
        return cfg
    except Exception:
        return {}


def _sync_cfg_from_widgets(cfg: dict) -> None:
    """Mirror the current widget state from session_state into cfg.

    On rerun, st.session_state[w_KEY] already holds the user's latest
    widget value before our script runs. Pulling these values into cfg
    BEFORE computing the ● diff markers and hard-error list is essential:
    without this, validation and markers lag by one rerun (they'd see
    the prior rerun's final cfg, not the widget state that just changed).
    Widgets still write back to cfg during render, but that second write
    is a no-op once sync has already happened.
    """

    def _get(wk, default=None):
        return st.session_state.get(wk, default)

    start = _get("w_ACADEMIC_DATE_START_STRING")
    if isinstance(start, date):
        cfg["ACADEMIC_DATE_START_STRING"] = start.isoformat()

    end = _get("w_ACADEMIC_DATE_END_STRING")
    if isinstance(end, date):
        cfg["ACADEMIC_DATE_END_STRING"] = end.isoformat()

    # PGY3 cutoff is gated by a checkbox. If the checkbox has ever
    # rendered we trust it; otherwise keep whatever cfg already has.
    if "cb_PGY3_CUTOFF_DATE" in st.session_state:
        if st.session_state["cb_PGY3_CUTOFF_DATE"]:
            dp = st.session_state.get("dp_PGY3_CUTOFF_DATE")
            if isinstance(dp, date):
                cfg["PGY3_CUTOFF_DATE"] = dp.isoformat()
        else:
            cfg["PGY3_CUTOFF_DATE"] = ""

    for key in (
        "SIMULATION_RUNS",
        "POST_CALL_DAYS",
        "MIN_SPACING_DAYS_STRONG",
        "MIN_SPACING_DAYS_MILD",
        "MAX_CALLS_IN_WINDOW",
        "ROLLING_WINDOW_DAYS",
        "MAX_DIFF_SOFT",
        "MAX_DIFF_HARD",
    ):
        v = _get(f"w_{key}")
        if v is not None:
            cfg[key] = int(v)

    for key in ("USE_COMPLETED_CALLS", "INTERN_BLOCK1_WEEKDAY_CALLS"):
        v = _get(f"w_{key}")
        if v is not None:
            cfg[key] = int(bool(v))

    nf = _get("w_NIGHT_FLOAT_ROTATION_NAME")
    if nf is not None:
        cfg["NIGHT_FLOAT_ROTATION_NAME"] = str(nf)

    for key in WEIGHT_KEYS:
        v = _get(f"w_{key}")
        if v is not None:
            cfg[key] = float(v)

    for key in ("PICK_CANDIDATE_RANK_ORDER", "MONTE_CARLO_SCORE_ORDER"):
        v = _get(f"w_{key}")
        if v is not None:
            cfg[key] = [ln.strip() for ln in str(v).splitlines() if ln.strip()]


def _diff_marker(key: str, defaults: dict) -> str:
    return " ●" if st.session_state.config.get(key) != defaults.get(key) else ""


def _label(key: str, text: str, defaults: dict) -> str:
    return f"{text}{_diff_marker(key, defaults)}"


def _render_errors(key: str, errors_by_key: dict[str, list[str]]) -> None:
    for msg in errors_by_key.get(key, []):
        st.error(msg)


def _date_input_nullable(
    label: str,
    key: str,
    value_raw: str,
    help_text: str,
    fallback_date: date,
) -> str:
    """Nullable date input. Returns ISO string, or '' if cleared.

    The checkbox gates the picker because st.date_input can't itself be
    "blank" once edited. `fallback_date` is used only on the very first
    render when cfg is empty and no prior dp_* session_state exists —
    once the user picks a date, Streamlit's session_state persists it
    across uncheck/recheck cycles.
    """
    current = _parse_date_or_none(value_raw)
    enabled = st.checkbox(
        f"Enable {label.lower()}",
        value=current is not None,
        key=f"cb_{key}",
        help=help_text,
    )
    if not enabled:
        return ""
    picked = st.date_input(
        label,
        value=current or fallback_date,
        key=f"dp_{key}",
        label_visibility="collapsed",
    )
    return picked.isoformat() if picked else ""


def _render_common_section(defaults: dict, errors: dict[str, list[str]]) -> None:
    st.markdown("**Common** — most users only edit these.")

    cfg = st.session_state.config

    start_raw = str(cfg.get("ACADEMIC_DATE_START_STRING", ""))
    start_current = _parse_date_or_none(start_raw) or date.today()
    start_picked = st.date_input(
        _label("ACADEMIC_DATE_START_STRING", "Academic year start", defaults),
        value=start_current,
        key="w_ACADEMIC_DATE_START_STRING",
        help="First day of the academic year. All scheduling begins from this date.",
    )
    cfg["ACADEMIC_DATE_START_STRING"] = start_picked.isoformat() if start_picked else ""
    _render_errors("ACADEMIC_DATE_START_STRING", errors)

    end_raw = str(cfg.get("ACADEMIC_DATE_END_STRING", ""))
    end_current = _parse_date_or_none(end_raw) or date.today()
    end_picked = st.date_input(
        _label("ACADEMIC_DATE_END_STRING", "Academic year end", defaults),
        value=end_current,
        key="w_ACADEMIC_DATE_END_STRING",
        help="Last day of the academic year (inclusive).",
    )
    cfg["ACADEMIC_DATE_END_STRING"] = end_picked.isoformat() if end_picked else ""
    _render_errors("ACADEMIC_DATE_END_STRING", errors)

    # Fallback for PGY3 cutoff when enabling from empty: 14 days before
    # academic end. Avoids defaulting to today (which is out of range).
    end_for_fallback = _parse_date_or_none(cfg.get("ACADEMIC_DATE_END_STRING"))
    pgy3_fallback = (
        end_for_fallback - timedelta(days=14) if end_for_fallback else date.today()
    )
    cfg["PGY3_CUTOFF_DATE"] = _date_input_nullable(
        _label("PGY3_CUTOFF_DATE", "PGY3 graduation cutoff", defaults),
        "PGY3_CUTOFF_DATE",
        str(cfg.get("PGY3_CUTOFF_DATE", "")),
        "PGY3s are excluded from call on this date and after. "
        "Disable to keep PGY3s eligible all year.",
        fallback_date=pgy3_fallback,
    )
    _render_errors("PGY3_CUTOFF_DATE", errors)

    cfg["SIMULATION_RUNS"] = int(
        st.number_input(
            _label("SIMULATION_RUNS", "Number of simulation runs", defaults),
            min_value=1,
            max_value=100000,
            value=int(cfg.get("SIMULATION_RUNS", 1000)),
            step=100,
            key="w_SIMULATION_RUNS",
            help="More runs = better schedule but slower. 1000 is usually sufficient.",
        )
    )
    _render_errors("SIMULATION_RUNS", errors)

    cfg["USE_COMPLETED_CALLS"] = int(
        st.toggle(
            _label("USE_COMPLETED_CALLS", "Partial-year mode (seed from completed calls)", defaults),
            value=bool(int(cfg.get("USE_COMPLETED_CALLS", 0))),
            key="w_USE_COMPLETED_CALLS",
            help="ON: seed from completed_calls.xlsx and generate from the next day. "
            "OFF: generate a full year from scratch.",
        )
    )
    _render_errors("USE_COMPLETED_CALLS", errors)

    cfg["INTERN_BLOCK1_WEEKDAY_CALLS"] = int(
        st.toggle(
            _label("INTERN_BLOCK1_WEEKDAY_CALLS", "Interns take weekday calls in Block 1", defaults),
            value=bool(int(cfg.get("INTERN_BLOCK1_WEEKDAY_CALLS", 1))),
            key="w_INTERN_BLOCK1_WEEKDAY_CALLS",
            help="ON: interns cover weekday and weekend calls in Block 1. "
            "OFF: weekday calls in Block 1 are covered by Night Float.",
        )
    )
    _render_errors("INTERN_BLOCK1_WEEKDAY_CALLS", errors)


def _render_advanced_section(defaults: dict, errors: dict[str, list[str]]) -> None:
    cfg = st.session_state.config

    def _num(key: str, label: str, help_text: str, step: int = 1, min_value: int = 0):
        cfg[key] = int(
            st.number_input(
                _label(key, label, defaults),
                min_value=min_value,
                value=int(cfg.get(key, 0)),
                step=step,
                key=f"w_{key}",
                help=help_text,
            )
        )
        _render_errors(key, errors)

    _num("POST_CALL_DAYS", "Post-call rest days",
         "Minimum days off after a call before another can be assigned. Hard constraint.")
    _num("MIN_SPACING_DAYS_STRONG", "Strong spacing threshold (days)",
         "Calls within this many days incur a strong penalty (soft — not hard-blocked).",
         min_value=1)
    _num("MIN_SPACING_DAYS_MILD", "Mild spacing threshold (days)",
         "Calls within this many days incur a mild penalty.", min_value=1)
    _num("MAX_CALLS_IN_WINDOW", "Max calls per rolling window",
         "Rolling cap: no more than this many calls per N days. Set to 0 to disable.")
    _num("ROLLING_WINDOW_DAYS", "Rolling window (days)",
         "Size of the rolling window for the cap above.", min_value=1)
    _num("MAX_DIFF_SOFT", "Soft fairness threshold",
         "Call-count gap that triggers the soft fairness flag in ranking.", min_value=1)
    _num("MAX_DIFF_HARD", "Hard fairness threshold",
         "Call-count gap that triggers the hard fairness flag.", min_value=1)

    cfg["NIGHT_FLOAT_ROTATION_NAME"] = st.text_input(
        _label("NIGHT_FLOAT_ROTATION_NAME", "Night Float rotation code", defaults),
        value=str(cfg.get("NIGHT_FLOAT_ROTATION_NAME", "NF")),
        key="w_NIGHT_FLOAT_ROTATION_NAME",
        help="The exact name used for Night Float in the flow sheet (e.g. 'NF').",
    )
    _render_errors("NIGHT_FLOAT_ROTATION_NAME", errors)


def _render_expert_section(defaults: dict, errors: dict[str, list[str]]) -> None:
    cfg = st.session_state.config

    st.warning(
        "⚠️ These values control how the scheduler ranks candidates. They have "
        "been tuned through Monte Carlo testing. Modifying them may produce "
        "worse schedules. Change only if you understand the weighted-score "
        "system — see README §Scoring."
    )

    def _weight(key: str, label: str, help_text: str):
        cfg[key] = float(
            st.number_input(
                _label(key, label, defaults),
                min_value=0.0,
                value=float(cfg.get(key, 0.0)),
                step=0.25,
                format="%.2f",
                key=f"w_{key}",
                help=help_text,
            )
        )
        _render_errors(key, errors)

    _weight("FAIRNESS_GAP_WEIGHT", "Fairness gap weight",
            "How strongly to prefer residents behind in call count.")
    _weight("SPACING_WEIGHT", "Spacing weight",
            "How strongly to prefer residents with longer spacing since last call.")
    _weight("AVOID_WEIGHT", "Avoid-rotation weight",
            "Penalty for assigning call while on an AVOID rotation.")
    _weight("YEAR_BIAS_WEIGHT", "Year-bias weight",
            "How strongly to front-load PGY3s and back-load PGY2s.")
    _weight("PACE_WEIGHT", "Pace weight",
            "Corrective: penalizes residents ahead of their expected call pace.")
    _weight("LOOKAHEAD_WEIGHT", "Lookahead weight",
            "Anticipatory: prefers residents with less remaining eligibility runway.")

    _render_rank_order(
        key="PICK_CANDIDATE_RANK_ORDER",
        label="Candidate rank order",
        help_text=(
            "Order in which candidate ranking criteria are applied "
            "(lexicographic, one per line). All items required."
        ),
        allowed=ALLOWED_PICK_RANK_ORDER,
        defaults=defaults,
        errors=errors,
    )
    _render_rank_order(
        key="MONTE_CARLO_SCORE_ORDER",
        label="Monte Carlo score order",
        help_text=(
            "Order in which schedule-level metrics are prioritized when selecting "
            "the best Monte Carlo run (one per line). All items required."
        ),
        allowed=ALLOWED_MC_SCORE_ORDER,
        defaults=defaults,
        errors=errors,
    )


def _render_rank_order(
    key: str,
    label: str,
    help_text: str,
    allowed: set[str],
    defaults: dict,
    errors: dict[str, list[str]],
) -> None:
    cfg = st.session_state.config
    current = cfg.get(key) or []
    st.caption(f"Required items: {', '.join(sorted(allowed))}")
    text = st.text_area(
        _label(key, label, defaults),
        value="\n".join(current),
        key=f"w_{key}",
        height=150,
        help=help_text,
    )
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    cfg[key] = lines
    _render_errors(key, errors)


# ---------------------------------------------------------------------------
# Settings — header (diff count, reset, save)
# ---------------------------------------------------------------------------


def _count_diffs(defaults: dict) -> int:
    return sum(
        1
        for key in BEHAVIOR_KEYS
        if st.session_state.config.get(key) != defaults.get(key)
    )


@st.dialog("Reset to defaults?")
def _confirm_reset() -> None:
    st.write("Discard all your changes and return to the saved defaults in config.yaml?")
    cols = st.columns(2)
    if cols[0].button("Yes, reset", type="primary", key="dlg_reset_yes"):
        fresh, _ = load_default_config()
        st.session_state.config = fresh
        # Clear per-widget keys so widgets re-initialise from the fresh config.
        for k in list(st.session_state.keys()):
            if k.startswith("w_") or k.startswith("cb_") or k.startswith("dp_"):
                del st.session_state[k]
        st.rerun()
    if cols[1].button("Cancel", key="dlg_reset_no"):
        st.rerun()


@st.dialog("Save as defaults?")
def _confirm_save(values: dict) -> None:
    st.write(
        "Overwrite config.yaml so your current values become the new defaults? "
        "The baseline will be permanently replaced. Comments in config.yaml are preserved."
    )
    cols = st.columns(2)
    if cols[0].button("Yes, save", type="primary", key="dlg_save_yes"):
        try:
            save_config(values)
            st.success("Defaults saved to config.yaml.")
        except Exception as exc:
            st.error(f"Failed to save: {exc}")
        st.rerun()
    if cols[1].button("Cancel", key="dlg_save_no"):
        st.rerun()


def _render_settings_header(defaults: dict, hard_errors: dict[str, list[str]]) -> None:
    cols = st.columns([4, 1, 1])
    diffs = _count_diffs(defaults)
    n_errors = sum(len(v) for v in hard_errors.values())
    msg_bits = []
    if diffs:
        msg_bits.append(f"{diffs} value(s) differ from defaults")
    if n_errors:
        msg_bits.append(f"**{n_errors} validation error(s) must be fixed**")
    cols[0].markdown(" · ".join(msg_bits) if msg_bits else "_All values match config.yaml defaults._")
    if cols[1].button("Reset to defaults", key="btn_reset"):
        _confirm_reset()
    # Only allow Save when there are no hard errors — saving an invalid
    # config would break the CLI too.
    save_disabled = n_errors > 0
    if cols[2].button("Save as defaults", key="btn_save", disabled=save_disabled):
        _confirm_save({k: st.session_state.config[k] for k in BEHAVIOR_KEYS})


def _render_settings_section() -> tuple[bool, list[str]]:
    """Render the whole Settings expander. Returns (hard_errors_exist, soft_warnings)."""
    defaults = _load_defaults_fresh()
    # Sync widget state → cfg BEFORE computing markers and errors so both
    # reflect the user's most recent interaction, not the prior rerun's
    # final cfg. Without this, markers and inline errors lag by one rerun.
    _sync_cfg_from_widgets(st.session_state.config)
    hard_errors = _compute_hard_errors(st.session_state.config)

    _render_settings_header(defaults, hard_errors)

    st.divider()
    with st.container():
        _render_common_section(defaults, hard_errors)
    with st.expander("Advanced", expanded=False):
        _render_advanced_section(defaults, hard_errors)
    with st.expander("Expert — tuned values, modify with caution", expanded=False):
        _render_expert_section(defaults, hard_errors)

    soft = _compute_soft_warnings(st.session_state.config)
    return bool(hard_errors), soft


# ---------------------------------------------------------------------------
# Page
# ---------------------------------------------------------------------------


def main() -> None:
    st.set_page_config(page_title="Call Schedule Creator", layout="wide")
    _init_session_state()

    header_cols = st.columns([6, 1])
    header_cols[0].title("Call Schedule Creator")
    if header_cols[1].button("Reset session"):
        _reset_session()
        st.rerun()

    with st.expander("1. Upload input files", expanded=True):
        for slot in UPLOAD_SLOTS:
            _render_upload_slot(slot)

    with st.expander("2. Settings", expanded=True):
        settings_has_errors, soft_warnings = _render_settings_section()

    with st.expander("3. Run schedule", expanded=False):
        uploads_ready = _all_required_valid()
        if not uploads_ready:
            st.warning("Upload all required files above before running.")
        if settings_has_errors:
            st.error("Fix the validation errors in Settings before running.")
        if soft_warnings:
            with st.expander(f"Configuration warnings ({len(soft_warnings)})", expanded=False):
                for w in soft_warnings:
                    st.warning(w)
        run_disabled = not uploads_ready or settings_has_errors
        st.button("Run (coming in Phase 5c)", disabled=run_disabled or True)

    with st.expander("4. Results", expanded=False):
        st.caption("Results will appear here after a run completes — Phase 5c.")


if __name__ == "__main__":
    main()
