from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple, Iterable
from openpyxl import load_workbook

from config import CONFIG

ROTATION_RULES_XLSX = CONFIG.get("ROTATION_RULES_XLSX", "data/rotation_rules.xlsx")
NO_CALL_DAYS_XLSX = CONFIG.get("NO_CALL_DAYS_XLSX", "data/no_call_days.xlsx")
HOLIDAYS_XLSX = CONFIG.get("HOLIDAYS_XLSX", "data/holidays.xlsx")
CLINIC_DAYS_XLSX = CONFIG.get("CLINIC_DAYS_XLSX", "data/clinic_days.xlsx")
COMPLETED_CALLS_XLSX = CONFIG.get("COMPLETED_CALLS_XLSX", "")


def load_residents(lookup) -> Dict[str, dict]:
    ws = lookup.ws
    skip_rows = lookup.skip_rows
    separator_pgy_labels: dict = getattr(lookup, "separator_pgy_labels", {})

    residents = {}
    pgy = 1
    separator_count = 0

    for r in range(3, ws.max_row + 1):
        if r in skip_rows:
            separator_count += 1
            explicit_pgy = separator_pgy_labels.get(r)
            if explicit_pgy is not None:
                pgy = explicit_pgy
            else:
                pgy += 1
                if pgy > 3:
                    raise ValueError(
                        f"Flow sheet has {separator_count} PGY separator rows, implying "
                        f"PGY{pgy} which is unexpected. Expected at most 2 separators "
                        f"(for PGY1/PGY2/PGY3). Check for extra header rows, or add "
                        f"explicit PGY labels (e.g. 'PGY2') in column A of separator rows."
                    )
            continue

        name = ws.cell(row=r, column=1).value
        if not name:
            continue

        name = str(name).strip()

        residents[name] = {
            "pgy": pgy,
            "assigned_dates": [],
            "total_calls": 0,
            "weekday_calls": 0,
            "weekend_calls": 0,
            "upper_calls": 0,
            "intern_calls": 0,
            "Jul_Dec_calls": 0,
            "Jan_Jun_calls": 0,
        }

    return residents


def _normalize_header(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def _parse_date(value) -> date:
    if value is None or value == "":
        raise ValueError("Blank date cell encountered while loading Excel input file.")

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    raise ValueError(f"Could not parse date value '{value}' from Excel input file.")


def _iter_excel_dict_rows(path: str) -> Iterable[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = [_normalize_header(cell.value) for cell in ws[1]]
    if not any(headers):
        return

    for r in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue
        yield {headers[i]: row_values[i] for i in range(len(headers))}


def load_rotation_rules(path: str = ROTATION_RULES_XLSX) -> Dict[Tuple[str, int], str]:
    rules: Dict[Tuple[str, int], str] = {}
    for row in _iter_excel_dict_rows(path):
        rotation = str(row["rotation_name"]).strip()
        pgy = int(row["pgy"])
        pref = str(row["preference"]).strip().upper()
        rules[(rotation, pgy)] = pref
    return rules


def load_no_call_days(path: str = NO_CALL_DAYS_XLSX) -> Dict[str, dict]:
    out: Dict[str, dict] = {}
    try:
        for row in _iter_excel_dict_rows(path):
            name = str(row["name"]).strip()
            start = _parse_date(row["start_date"])
            end = _parse_date(row["end_date"])
            reason = str(row.get("type", "") or "").strip()

            cur = start
            while cur <= end:
                out.setdefault(name, {})[cur] = reason
                cur += timedelta(days=1)
    except FileNotFoundError:
        pass

    return out


def load_holidays(path: str = HOLIDAYS_XLSX) -> Dict[date, str]:
    out: Dict[date, str] = {}
    try:
        for row in _iter_excel_dict_rows(path):
            d = _parse_date(row["date"])
            out[d] = str(row.get("name", "") or "").strip() or "Holiday"
    except FileNotFoundError:
        pass
    return out


def load_clinic_days(path: str = CLINIC_DAYS_XLSX) -> Dict[str, Dict[date, str]]:
    """Load clinic days and return the day BEFORE each clinic as a no-call day.

    Residency rule: a resident cannot take call the night before a clinic
    because the following morning they must be present in clinic (post-call
    rest day would conflict). The day before each clinic date is therefore
    blocked with reason 'pre_clinic_day'.

    File format: columns 'name' and 'date' (one row per clinic day).
    If the file does not exist, returns an empty dict (clinics are optional).
    """
    out: Dict[str, Dict[date, str]] = {}
    try:
        for row in _iter_excel_dict_rows(path):
            name = str(row["name"]).strip()
            clinic_date = _parse_date(row["date"])
            blocked_date = clinic_date - timedelta(days=1)
            out.setdefault(name, {})[blocked_date] = "pre_clinic_day"
    except FileNotFoundError:
        pass
    return out


def load_completed_calls(path: str = COMPLETED_CALLS_XLSX, block1_end: Optional[date] = None) -> List[Tuple[date, str, str]]:
    """Load a partially-completed call schedule to seed a mid-year restart.

    File format: one row per call day, with columns:
      - a 'date' column (any header containing 'date')
      - an upper-level column (any header containing 'upper')
      - an intern column (any header containing 'intern')
    Blank cells mean no assignment for that slot on that date.

    Slot types are inferred from the date: weekday → UPPER_WEEKDAY,
    weekend upper → UPPER_WEEKEND, weekend intern → INTERN_WEEKEND.

    Returns a list of (date, resident_name, slot) tuples sorted by date.
    Returns an empty list if the file does not exist or has no data.
    """
    if not path:
        return []

    out: List[Tuple[date, str, str]] = []
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        headers = [_normalize_header(cell.value) for cell in ws[1]]
        if not any(headers):
            return []

        date_col: Optional[int] = next(
            (i for i, h in enumerate(headers) if "date" in h), None
        )
        upper_col: Optional[int] = next(
            (i for i, h in enumerate(headers) if "upper" in h), None
        )
        intern_col: Optional[int] = next(
            (i for i, h in enumerate(headers) if "intern" in h), None
        )

        if date_col is None:
            raise ValueError(
                f"No column containing 'date' found in {path}. "
                f"Headers found: {headers}"
            )

        for r in range(2, ws.max_row + 1):
            row_vals = [
                ws.cell(row=r, column=c + 1).value for c in range(len(headers))
            ]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue

            try:
                d = _parse_date(row_vals[date_col])
            except (ValueError, TypeError):
                continue

            if upper_col is not None:
                upper_name = row_vals[upper_col]
                if upper_name and str(upper_name).strip():
                    slot = "UPPER_WEEKEND" if d.weekday() >= 5 else "UPPER_WEEKDAY"
                    out.append((d, str(upper_name).strip(), slot))

            if intern_col is not None:
                intern_name = row_vals[intern_col]
                if intern_name and str(intern_name).strip():
                    if d.weekday() >= 5:
                        out.append((d, str(intern_name).strip(), "INTERN_WEEKEND"))
                    elif block1_end is not None and d <= block1_end:
                        # Block 1 intern weekday call (feature enabled)
                        out.append((d, str(intern_name).strip(), "INTERN_WEEKDAY"))
                    # else: weekday intern = Night Float — silently dropped

    except FileNotFoundError:
        pass

    return sorted(out, key=lambda x: x[0])
